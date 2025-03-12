from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, ElementNotInteractableException
from selenium.common.exceptions import JavascriptException, NoSuchElementException
from datetime import datetime
import openpyxl
import time
import tkinter as tk
from tkinter import messagebox
import subprocess
import threading
import queue

def show_success_popup():
    root = tk.Tk()
    root.withdraw()  # Esconde a janela principal
    messagebox.showinfo("Sucesso", "A operação foi executada com sucesso!")
    root.destroy()

def execute_javascript_with_retry(driver, script, max_attempts=3, refresh_delay=10, retry_delay=5):
    attempts = 0
    while attempts < max_attempts:
        try:
            driver.execute_script(script)
            print("Script JavaScript executado com sucesso.")
            return True
        except JavascriptException as js_error:
            print(f"Erro ao executar script JavaScript: {js_error}. Tentando novamente em {retry_delay} segundos...")
            time.sleep(retry_delay)
        except NoSuchElementException:
            print("Elemento não encontrado, mas continuando a execução.")
            return False
        except Exception as e:
            print(f"Ocorreu um erro: {e}. Atualizando a página e repetindo operação em {refresh_delay} segundos.")
            driver.refresh()
            time.sleep(refresh_delay)

        attempts += 1

    print(f"Falha ao executar o script após {max_attempts} tentativas.")
    return False


def verificar_erro_processamento(driver):
    script = """
        return (document.querySelector('h2') || {}).textContent.trim() === 'Erro no processamento da solicitação';
    """
    try:
        if driver.execute_script(script):
            print("Erro no processamento da solicitação. Atualizando a página e repetindo operação em 10 segundos.")
            driver.get("https://geridinss.dataprev.gov.br/gpa")
            success = execute_javascript_with_retry(driver, "document.getElementById('formMenu:btAtribAcesso').click();")
            if success:
                print("O botão foi clicado com sucesso.")
            else:
                print("Não foi possível clicar no botão após várias tentativas.")
            return True
        return False
    except Exception as e:
        print(f"Erro ao verificar elemento: {e}")
        return False



def verificar_mensagem_operacao(driver, worksheet, linha, coluna):
    mensagens = {
        "/html/body/div[1]/div[2]/form[1]/div[2]/ul/li": [
            "A operação foi executada com sucesso."
        ],
        "/html/body/div[1]/div[2]/ul/li": [
            "Domínio não existe.",
            "Ocorreu um erro de comunicação. Aguarde alguns minutos e tente novamente.",
            "A Data de Validade não deve ser superior a Data de Validade da credencial do usuário emissor.",
            "Gestor de Acesso só pode atribuir acesso no seu próprio domínio ou domínio abaixo de sua abrangência.",
            "Não é permitido dar uma autorização a si mesmo."
        ]
    }

    for xpath, textos_esperados in mensagens.items():
        try:
            elements = driver.find_elements(By.XPATH, xpath)
            if len(elements) == 1 and elements[0].text in textos_esperados:
                mensagem = elements[0].text
                print(mensagem)
                worksheet.cell(row=linha, column=coluna + 4).value = mensagem
                return True, coluna + 5
        except:
            continue
    
    return False, coluna




def run_automation(file_path, update_label_func=None, update_status_func=None):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    Finaliza = True

    try:
        linha = 2
        coluna = 3

        driver.get("https://geridinss.dataprev.gov.br/gpa")
        driver.implicitly_wait(10)
        driver.maximize_window()

        print(f"Aguardando procedimento de Login...")
        WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/ul/li")))
        print(f"Login realizado com sucesso.")

        success = execute_javascript_with_retry(driver, "document.getElementById('formMenu:btAtribAcesso').click();")
        if success:
            print("O botão foi clicado com sucesso.")
        else:
            print("Não foi possível clicar no botão após várias tentativas.")

        
        if update_status_func:
            update_status_func("Processando registros...")
        
        
        while Finaliza:
            try:
                servidor = worksheet.cell(row=linha, column=1).value
                Sistema = worksheet.cell(row=linha, column=coluna).value
                Subsistema = worksheet.cell(row=linha, column=coluna + 1).value
                Papel = worksheet.cell(row=linha, column=coluna + 2).value
                UO = worksheet.cell(row=linha, column=2).value
                validade = worksheet.cell(row=linha, column=coluna + 3).value
                Situacao = worksheet.cell(row=linha, column=coluna + 4).value

                if not servidor:
                    try:
                        workbook.save(file_path)
                    except Exception as e:
                        print(f"Erro ao salvar o arquivo: {e}")
                    driver.quit()
                    print("Final!")
                    show_success_popup()
                    break

                if not Sistema:
                    coluna = 3
                    linha += 1

                    if update_label_func:
                        update_label_func(linha)

                    continue

                if Situacao is not None:
                    coluna += 5
                    continue

                print("===================================================================")
                print(f"Acesso: {servidor} - {UO} - {Sistema} - {Subsistema} - {Papel} - {validade}")




                # # Verificar se ocorreu o erro específico================
                # error_elements = driver.find_elements(By.CSS_SELECTOR, "#form > fieldset > div:nth-child(2) > label > span")
                # if error_elements and error_elements[0].text == "Erro no processamento da solicitação":
                #     print("Erro no processamento da solicitação. Atualizando a página e repetindo operação em 10 segundos.")
                #     driver.get("https://geridinss.dataprev.gov.br/gpa")
                #     success = execute_javascript_with_retry(driver, "document.getElementById('formMenu:btAtribAcesso').click();")
                #     if success:
                #         print("O botão foi clicado com sucesso.")
                #     else:
                #         print("Não foi possível clicar no botão após várias tentativas.")
                #     continue
                # # Verificar se ocorreu o erro específico================
                
                #a lógica abaixo verifica se o sistema está disponível e se está na pagina correta para consulta do SISTEMA
                sistema_label_xpath = "/html/body/div[1]/div[2]/form[1]/fieldset/div[1]/label/span"
                try:
                    sistema_label = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, sistema_label_xpath)))
                    try:
                        Select(driver.find_element(By.ID, "form:sistema")).select_by_visible_text(Sistema)
                    except (NoSuchElementException, ElementNotInteractableException):
                        print("Sistema não localizado dentre as opções disponíveis.")
                        worksheet.cell(row=linha, column=coluna + 4).value = "Sistema não localizado dentre as opções disponíveis."
                        coluna += 5
                        continue
                except TimeoutException:
                    driver.get("https://geridinss.dataprev.gov.br/gpa")
                    success = execute_javascript_with_retry(driver, "document.getElementById('formMenu:btAtribAcesso').click();")
                    if success:
                        pass
                    else:
                        print("Não foi possível clicar no botão após várias tentativas.")
                    continue
                #a lógica abaixo verifica se o sistema está disponível e se está na pagina correta para consulta do SISTEMA

                try:
                    Select(driver.find_element(By.ID, "form:subsistema")).select_by_visible_text(Subsistema)
                    Select(driver.find_element(By.ID, "form:papel")).select_by_visible_text(Papel)
                    Select(driver.find_element(By.ID, "form:tipoDominio")).select_by_visible_text("UO")
                    driver.find_element(By.ID, "form:dominio").clear()
                    driver.find_element(By.ID, "form:dominio").send_keys(UO)
                    driver.find_element(By.ID, "form:usuario").clear()
                    driver.find_element(By.ID, "form:usuario").send_keys(servidor)
                    driver.find_element(By.ID, "form:filtrar").click()
                except (NoSuchElementException, ElementNotInteractableException) as e:
                    print(f"Erro ao interagir com elementos da página: {e}")
                    print("Atualizando a página e tentando novamente...")
                    driver.get("https://geridinss.dataprev.gov.br/gpa")
                    time.sleep(5)
                    continue
                
                #verificando erro de comunicação====================================================================
                xpath = "/html/body/div[1]/div[2]/ul/li"
                try:                    
                    element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
                    print("Erro de comunicação detectado. Aguardando e tentando novamente...")
                    driver.get("https://geridinss.dataprev.gov.br/gpa")
                    execute_javascript_with_retry(driver, "document.getElementById('formMenu:btAtribAcesso').click();") 
                    continue
                    
                except Exception as e:
                    pass
                print("Nenhum erro de comunicação detectado.")
                #verificando erro de comunicação====================================================================
                
                

                # APTO PARA REVALIDAÇÃO=============================================================================
                try:
                    element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/form[2]/table/tbody/tr/td[7]")))
                    data_string = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/form[2]/table/tbody/tr/td[7]").text
                    data = datetime.strptime(data_string, "%d/%m/%Y")
                    validade_formatada = validade.strftime("%d/%m/%Y")
                    data_formatada = data.strftime("%d/%m/%Y")
                    print("Data de validade informada:", validade_formatada)
                    print("Data de validade atual:", data_formatada)
                    data_string_dt = datetime.strptime(data_string, "%d/%m/%Y")

                    try:
                        workbook.save(file_path)
                    except Exception as e:
                        print(f"Erro ao salvar o arquivo: {e}")

                    if data_string_dt < validade:
                        print("A data extraída do gerid é menor que a data informada. Avançando para revalidação...")
                        driver.find_element(By.ID, "dataTableCredencial:selected").click()
                        driver.find_element(By.ID, "form2:btAlterar").click()
                        driver.find_element(By.ID, "form2:dataValidade").send_keys(Keys.CONTROL, 'a')
                        validade_str = validade.strftime("%d/%m/%Y")
                        driver.find_element(By.ID, "form2:dataValidade").send_keys(validade_str)
                        driver.find_element(By.ID, "form2:confirmar").click()

                        # VERIFICA SE DEU CERTO!====================:
                        sucesso, nova_coluna = verificar_mensagem_operacao(driver, worksheet, linha, coluna)
                        if sucesso:
                            coluna = nova_coluna
                            continue
                        # VERIFICA SE DEU CERTO!====================:
                        
                        
                        
                        
                        
                        
                        

                        # salvando==============================
                        try:
                            workbook.save(file_path)
                        except Exception as e:
                            print(f"Erro ao salvar o arquivo: {e}")
                        # salvando==============================

                    else:
                        worksheet.cell(row=linha, column=coluna + 4).value = f"Sistema já revalidado ({validade.strftime('%d/%m/%Y')})"
                        print(f"Sistema já revalidado ({validade.strftime('%d/%m/%Y')})")
                        # salvando==============================
                        try:
                            workbook.save(file_path)
                        except Exception as e:
                            print(f"Erro ao salvar o arquivo: {e}")
                        # salvando==============================

                except (NoSuchElementException, TimeoutException):
                    print("Atribuindo novo acesso....")
                    try:
                        driver.find_element(By.ID, "form2:novo").click()
                        Select(driver.find_element(By.ID, "form:sistema")).select_by_visible_text(Sistema)
                        Select(driver.find_element(By.ID, "form:subsistema")).select_by_visible_text(Subsistema)
                        Select(driver.find_element(By.ID, "form2:papel")).select_by_visible_text(Papel)
                        Select(driver.find_element(By.ID, "form2:tipoDominio")).select_by_visible_text("UO")
                        driver.find_element(By.ID, "form2:dominio").clear()
                        driver.find_element(By.ID, "form2:dominio").send_keys(UO)
                        driver.find_element(By.ID, "form2:usuario").click()
                        driver.find_element(By.ID, "form2:usuario").clear()
                        driver.find_element(By.ID, "form2:usuario").send_keys(servidor)
                        driver.find_element(By.ID, "form2:dominio").click()
                        driver.find_element(By.ID, "form2:dataValidade").click()
                        driver.find_element(By.ID, "form2:dataValidade").clear()
                        driver.find_element(By.ID, "form2:dataValidade").send_keys(Keys.CONTROL, 'a')

                        # Convertendo o objeto datetime para string no formato correto
                        validade_str = validade.strftime("%d/%m/%Y")
                        driver.find_element(By.ID, "form2:dataValidade").send_keys(validade_str)

                        # Selecionando os dias da semana
                        driver.find_element(By.ID, "form2:periodo:0").click()
                        driver.find_element(By.ID, "form2:periodo:6").click()
                        driver.find_element(By.ID, "form2:periodo:7").click()

                        # Configurando horários
                        driver.find_element(By.ID, "form2:horaAcessoInicio").click()
                        driver.find_element(By.ID, "form2:horaAcessoInicio").send_keys(Keys.CONTROL, 'a')
                        driver.find_element(By.ID, "form2:horaAcessoInicio").send_keys("0000")

                        driver.find_element(By.ID, "form2:horaAcessoFim").click()
                        driver.find_element(By.ID, "form2:horaAcessoFim").send_keys(Keys.CONTROL, 'a')
                        driver.find_element(By.ID, "form2:horaAcessoFim").send_keys("2359")

                        driver.find_element(By.ID, "form2:confirmar").click()

                        # Verificando mensagens de retorno
                        try:
                            element = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/ul/li")
                            mensagem = element.text
                            print(mensagem)
                            worksheet.cell(row=linha, column=coluna + 4).value = mensagem
                            coluna += 5
                            continue
                        except:
                            pass

                        print(f"Acesso: {UO} - {Sistema} - {Subsistema} - {Papel} - {servidor} - {validade_str}")

                        # Aguardando e verificando mensagem de sucesso
                        wait = WebDriverWait(driver, 3)
                        element = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/form[1]/div[2]/ul/li")))
                        mensagem_sucesso = element.text
                        print(mensagem_sucesso)
                        worksheet.cell(row=linha, column=coluna + 4).value = mensagem_sucesso
                        print()

                        # Salvando a planilha
                        try:
                            workbook.save(file_path)
                        except Exception as e:
                            print(f"Erro ao salvar o arquivo: {e}")

                        coluna += 5
                        continue

                    except Exception as e:
                        print(f"Erro ao atribuir novo acesso repetindo operação: {e}")
                        # worksheet.cell(row=linha, column=coluna + 4).value = f"Erro ao atribuir novo acesso: {str(e)}"
                        # workbook.save(file_path)
                        # coluna += 5
                        driver.get("https://geridinss.dataprev.gov.br/gpa")
                        success = execute_javascript_with_retry(driver, "document.getElementById('formMenu:btAtribAcesso').click();")
                        if success:
                            print("O botão foi clicado com sucesso.")
                        else:
                            print("Não foi possível clicar no botão após várias tentativas.")
                        continue
            except Exception as e:
                print(f"Erro inesperado: {e}")
                print("Tentando recuperar e continuar...")
                driver.get("https://geridinss.dataprev.gov.br/gpa")
                success = execute_javascript_with_retry(driver, "document.getElementById('formMenu:btAtribAcesso').click();")
                if success:
                    print("O botão foi clicado com sucesso.")
                else:
                    print("Não foi possível clicar no botão após várias tentativas.")

                time.sleep(5)
                continue

        print("==== FIM ====")
    except Exception as e:
        print(f"Erro geral: {e}")
        if update_status_func:
            update_status_func("Erro durante a execução...")
        
    finally:
        driver.quit()
        print("Driver encerrado com sucesso.")

    if update_status_func:
        update_status_func("Execução finalizada.")

def run_automation_thread(file_path, update_label_func=None, update_status_func=None):
    thread = threading.Thread(target=run_automation, args=(file_path, update_label_func, update_status_func))
    thread.start()
    return thread  # Retorna a thread para que possamos acompanhar seu estado
