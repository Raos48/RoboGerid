from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, ElementNotInteractableException
from selenium.common.exceptions import JavascriptException, NoSuchElementException
from datetime import datetime
import openpyxl
import time
import tkinter as tk
from tkinter import messagebox

file_path = "C:\\Users\\Ricardo\\OneDrive - INSS\\PYTHON\\Bot SAGGESTAO GERID\\MODELO ATRIBUIÇÃO E REVALIDAÇÃO ACESSOS NORMAIS 08021.xlsx"

workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active

service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)
Finaliza = True

linha = 2
coluna = 3

driver.get("https://geridinss.dataprev.gov.br/gpa")
driver.implicitly_wait(10)
driver.maximize_window()

print(f"Aguardando procedimento de Login...")
WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/ul/li")))
print(f"Login realizado com sucesso.")

# CLICANDO NO ELEMENTO ATRIBUIR ACESSO=======================================
success = False
attempts = 0
while attempts < 3 and not success:
    try:
        driver.execute_script("document.getElementById('formMenu:btAtribAcesso').click();")
        print("O botão foi clicado com sucesso.")
        success = True
    except JavascriptException as js_error:
        print(f"Erro ao executar script JavaScript: {js_error}. Tentando novamente em 5 segundos...")
        time.sleep(5)
    except NoSuchElementException:
        print("Elemento não encontrado, mas continuando a execução.")
        success = True
    except Exception as e:
        print(f"Ocorreu um erro: {e}. Atualizando a página e repetindo operação em 10 segundos.")
        driver.refresh()
        time.sleep(10)
    attempts += 1

if not success:
    print(f"Falha ao executar o script após 3 tentativas.")
# CLICANDO NO ELEMENTO ATRIBUIR ACESSO=======================================






def show_success_popup():
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo("Sucesso", "A operação foi executada com sucesso!")
    root.destroy()





def realizar_pesquisa(driver, Sistema, Subsistema, Papel, UO, servidor, worksheet, linha):
    revalidar_acesso = False
    conceder_novo_acesso = False

    while True:
        verificar_erros = False

        # CHECAGEM SE PESQUISA está visível===================================================
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#form > fieldset")))
            #print("✓ Pesquisa disponível.")
        except Exception as e:
            print(f"✗ ERRO Pesquisa não disponível.")
            print("Recarregando a página...")
            driver.get("https://geridinss.dataprev.gov.br/gpa")
            time.sleep(5)
            driver.execute_script("document.getElementById('formMenu:btAtribAcesso').click();")
            time.sleep(5)
        # CHECAGEM SE PESQUISA está visível===================================================

        # CHECAGEM SE SISTEMA está visível===================================================
        try:
            sistema_label_xpath = "/html/body/div[1]/div[2]/form[1]/fieldset/div[1]/label/span"
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, sistema_label_xpath)))
            #print("Campo sistema localizado")
        except Exception as e:
            print("Erro na verificação select SISTEMAS, atualizando página e tentando novamente...")
            driver.get("https://geridinss.dataprev.gov.br/gpa")
            driver.execute_script("document.getElementById('formMenu:btAtribAcesso').click();")
            time.sleep(5)
            continue
        # CHECAGEM SE SISTEMA está visível===================================================

        # OBTER LISTA DE SISTEMAS DISPONÍVEIS DEPOIS COMPARAR COM SISTEMA SELECIONADO=============
        sistemas_disponiveis = [sistema.strip() for sistema in driver.find_element(By.ID, "form:sistema").text.split('\n')]
        sistemas_disponiveis = sistemas_disponiveis[1:]
        if Sistema.strip() in sistemas_disponiveis:
            print("✓ SISTEMA/SUBSISTEMA disponível")
        else:
            print("✗ ERRO: Sistema não localizado")
            print(f"Sistema '{Sistema}' não está entre as opções: {sistemas_disponiveis}")
            worksheet.cell(row=linha, column=7).value = "Sistema não localizado dentre as opções disponíveis."
            print("Avançando para próxima iteração...")
            linha+=1
            continue
        # OBTER LISTA DE SISTEMAS DISPONÍVEIS DEPOIS COMPARAR COM SISTEMA SELECIONADO=============

        # PREENCHIMENTO DOS CAMPOS=================================================================
        while True:
            try:
                Select(driver.find_element(By.ID, "form:sistema")).select_by_visible_text(Sistema)
                Select(driver.find_element(By.ID, "form:subsistema")).select_by_visible_text(Subsistema)
                Select(driver.find_element(By.ID, "form:papel")).select_by_visible_text(Papel)
                Select(driver.find_element(By.ID, "form:tipoDominio")).select_by_visible_text("UO")
                driver.find_element(By.ID, "form:dominio").clear()
                driver.find_element(By.ID, "form:dominio").send_keys(UO)
                driver.find_element(By.ID, "form:usuario").clear()
                driver.find_element(By.ID, "form:usuario").send_keys(servidor)
                driver.find_element(By.ID, "form:filtrar").click()
                break
            except Exception as e:
                print("✗ ERRO ao tentar preencher informações, atualizando página e repetindo processo..")
                driver.get("https://geridinss.dataprev.gov.br/gpa")
                driver.execute_script("document.getElementById('formMenu:btAtribAcesso').click();")
                time.sleep(5)
                continue
        # PREENCHIMENTO DOS CAMPOS=================================================================

        # VERIFICANDO SE A TABELA FOI RETORNADA COM SUCESSO========================================
        try:
            tabela_xpath = "/html/body/div[1]/div[2]/form[2]/table/tbody"
            tabela = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, tabela_xpath)))
            primeiro_campo = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/form[2]/table/tbody/tr/td[2]").text
            if primeiro_campo:
                print("Tabela de Acessos.")
                verificar_erros = False
                element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/form[2]/table/tbody/tr/td[7]")))
                data_string = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/form[2]/table/tbody/tr/td[7]").text
                data = datetime.strptime(data_string, "%d/%m/%Y")
                validade_formatada = validade.strftime("%d/%m/%Y")
                data_formatada = data.strftime("%d/%m/%Y")
                print("Data de validade informada:", validade_formatada)
                print("Data de validade atual:", data_formatada)
                data_string_dt = datetime.strptime(data_string, "%d/%m/%Y")
                if data_string_dt < validade:
                    revalidar_acesso = True
                else:
                    revalidar_acesso = False
                break
        except (NoSuchElementException, TimeoutException) as e:
            #print("Tabela não retornada ou erro ao localizar elemento.")
            verificar_erros = True
        # VERIFICANDO SE A TABELA FOI RETORNADA COM SUCESSO==========================================

        # VERIFICANDO ERROS==================================================
        if verificar_erros:
            # ERRO DE COMUNICAÇÃO XXXXXXXXXXXXXXXXXXXXXXX
            try:
                erro_comunicacao = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/form[1]/div[2]/ul/li").text
                if erro_comunicacao == 'Ocorreu um erro de comunicação. Aguarde alguns minutos e tente novamente.':
                    print("✗ ERRO de comunicação detectado.")
                    print("Atualizando a página e tentando novamente...")
                    driver.get("https://geridinss.dataprev.gov.br/gpa")
                    driver.execute_script("document.getElementById('formMenu:btAtribAcesso').click();")
                    time.sleep(5)
                    continue
            except Exception as e:
                print("Nenhum erro de comunicação detectado.")

            # ERRO REGRA NEGÓCIO XXXXXXXXXXXXXXXXXXX
            try:
                erro_negocio = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/form[1]/div[2]/ul/li").text
                if erro_negocio == 'Erro ao executar regras de negócio':
                    print("✗ ERRO de regras de negócio detectado.")
                    print("Atualizando a página e tentando novamente...")
                    driver.get("https://geridinss.dataprev.gov.br/gpa")
                    driver.execute_script("document.getElementById('formMenu:btAtribAcesso').click();")
                    time.sleep(5)
                    continue
            except Exception as e:
                print("Nenhum erro de negócio detectado.")
            conceder_novo_acesso = True
        break

    return revalidar_acesso, conceder_novo_acesso

def revalida_acesso(driver, validade, worksheet, linha, workbook, file_path):
    #print("A data extraída do gerid é menor que a data informada. Avançando para revalidação...")
    driver.find_element(By.ID, "dataTableCredencial:selected").click()
    driver.find_element(By.ID, "form2:btAlterar").click()
    driver.find_element(By.ID, "form2:dataValidade").send_keys(Keys.CONTROL, 'a')
    validade_str = validade.strftime("%d/%m/%Y")
    driver.find_element(By.ID, "form2:dataValidade").send_keys(validade_str)
    driver.find_element(By.ID, "form2:confirmar").click()

    mensagens = [
        ("/html/body/div[1]/div[2]/form[1]/div[2]/ul/li", 'A operação foi executada com sucesso.'),
        ("/html/body/div[1]/div[2]/ul/li", 'Domínio não existe.'),
        ("/html/body/div[1]/div[2]/ul/li", 'Ocorreu um erro de comunicação. Aguarde alguns minutos e tente novamente.'),
        ("/html/body/div[1]/div[2]/ul/li", 'A Data de Validade não deve ser superior a Data de Validade da credencial do usuário emissor.'),
        ("/html/body/div[1]/div[2]/ul/li", 'Gestor de Acesso só pode atribuir acesso no seu próprio domínio ou domínio abaixo de sua abrangência.'),
        ("/html/body/div[1]/div[2]/ul/li", 'Não é permitido dar uma autorização a si mesmo.')
    ]

    for xpath, texto in mensagens:
        try:
            wait = WebDriverWait(driver, 3)
            element = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
            mensagem = element.text
            if mensagem == texto:
                print(f"{'✓' if texto == 'A operação foi executada com sucesso.' else '✗'} {mensagem}")
                worksheet.cell(row=linha, column=7).value = mensagem
                _salvar_planilha(workbook, file_path)
                return (texto == 'A operação foi executada com sucesso.'), mensagem
        except NoSuchElementException:
            continue

    worksheet.cell(row=linha, column=7).value = "Erro não identificado"
    _salvar_planilha(workbook, file_path)
    return False, "Erro não identificado"

# Função auxiliar para salvar planilha
def _salvar_planilha(workbook, file_path):
    try:
        workbook.save(file_path)
    except Exception as e:
        print(f"Erro ao salvar o arquivo: {e}")

def atribuir_novo_acesso(driver, validade):
    try:
        driver.find_element(By.ID, "form2:novo").click()

        Select(driver.find_element(By.ID, "form:sistema")).select_by_visible_text(Sistema)
        Select(driver.find_element(By.ID, "form:subsistema")).select_by_visible_text(Subsistema)
        Select(driver.find_element(By.ID, "form2:papel")).select_by_visible_text(Papel)
        Select(driver.find_element(By.ID, "form2:tipoDominio")).select_by_visible_text("UO")
        _limpar_e_enviar_keys(driver, "form2:dominio", UO)
        driver.find_element(By.ID, "form2:usuario").click()
        _limpar_e_enviar_keys(driver, "form2:usuario", servidor)
        driver.find_element(By.ID, "form2:dominio").click()
        driver.find_element(By.ID, "form2:dataValidade").click()
        _limpar_e_enviar_keys(driver, "form2:dataValidade", Keys.CONTROL + 'a')
        validade_str = validade.strftime("%d/%m/%Y")
        driver.find_element(By.ID, "form2:dataValidade").send_keys(validade_str)
        driver.find_element(By.ID, "form2:periodo:0").click()
        driver.find_element(By.ID, "form2:periodo:6").click()
        driver.find_element(By.ID, "form2:periodo:7").click()
        driver.find_element(By.ID, "form2:horaAcessoInicio").click()
        _limpar_e_enviar_keys(driver, "form2:horaAcessoInicio", Keys.CONTROL + 'a')
        driver.find_element(By.ID, "form2:horaAcessoInicio").send_keys("0000")
        driver.find_element(By.ID, "form2:horaAcessoFim").click()
        _limpar_e_enviar_keys(driver, "form2:horaAcessoFim", Keys.CONTROL + 'a')
        driver.find_element(By.ID, "form2:horaAcessoFim").send_keys("2359")
        driver.find_element(By.ID, "form2:confirmar").click()

        mensagens = [
            ("/html/body/div[1]/div[2]/form[1]/div[2]/ul/li", 'A operação foi executada com sucesso.'),
            ("/html/body/div[1]/div[2]/ul/li", 'Domínio não existe.'),
            ("/html/body/div[1]/div[2]/ul/li", 'Ocorreu um erro de comunicação. Aguarde alguns minutos e tente novamente.'),
            ("/html/body/div[1]/div[2]/ul/li", 'A Data de Validade não deve ser superior a Data de Validade da credencial do usuário emissor.'),
            ("/html/body/div[1]/div[2]/ul/li", 'Gestor de Acesso só pode atribuir acesso no seu próprio domínio ou domínio abaixo de sua abrangência.'),
            ("/html/body/div[1]/div[2]/ul/li", 'Não é permitido dar uma autorização a si mesmo.')
        ]

        for xpath, texto in mensagens:
            try:
                wait = WebDriverWait(driver, 3)
                element = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
                mensagem = element.text
                if mensagem == texto:
                    print(f"{'✓' if texto == 'A operação foi executada com sucesso.' else '✗'} {mensagem}")
                    worksheet.cell(row=linha, column=7).value = mensagem
                    print(f"Acesso: {UO} - {Sistema} - {Subsistema} - {Papel} - {servidor} - {validade_str}")
                    _salvar_planilha(workbook, file_path)
                    return (texto == 'A operação foi executada com sucesso.'), mensagem
            except NoSuchElementException:
                continue

        print("Mensagem de retorno não reconhecida.")
        worksheet.cell(row=linha, column=7).value = "Mensagem de retorno não reconhecida."
        _salvar_planilha(workbook, file_path)
        return False, "Mensagem de retorno não reconhecida."

    except Exception as e:
        print(f"Erro inesperado: {e}")
        worksheet.cell(row=linha, column=7).value = f"Erro inesperado: {e}"
        _salvar_planilha(workbook, file_path)
        return False, f"Erro inesperado: {e}"

def _limpar_e_enviar_keys(driver, element_id, keys):
    """Limpa o campo de texto e envia as chaves."""
    element = driver.find_element(By.ID, element_id)
    element.clear()
    element.send_keys(keys)

# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> CÓDIGO PRINCIPAL <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<
# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> CÓDIGO PRINCIPAL <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<
# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> CÓDIGO PRINCIPAL <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

while True:
    driver.get("https://geridinss.dataprev.gov.br/gpa")
    driver.execute_script("document.getElementById('formMenu:btAtribAcesso').click();")

    print("============================================================")

    servidor = worksheet.cell(row=linha, column=1).value
    UO = worksheet.cell(row=linha, column=2).value
    Sistema = worksheet.cell(row=linha, column=3).value
    Subsistema = worksheet.cell(row=linha, column=4).value
    Papel = worksheet.cell(row=linha, column=5).value
    validade = worksheet.cell(row=linha, column=6).value
    Situacao = worksheet.cell(row=linha, column=7).value

    print(f"{linha}-{servidor}-{UO}-{Sistema}/{Subsistema}/{Papel}-{validade}-Status:{Situacao}")

    if not servidor:
        try:
            workbook.save(file_path)
        except Exception as e:
            print(f"Erro ao salvar o arquivo: {e}")
        driver.quit()
        print("Final!")
        break

    if Situacao is not None:
        linha += 1
        continue

    # REALIZAR PESQUISA =====================================================================================================
    revalidar_acesso, conceder_novo_acesso = realizar_pesquisa(driver, Sistema, Subsistema, Papel, UO, servidor, worksheet, linha)

    print(revalidar_acesso, conceder_novo_acesso)
    
    if revalidar_acesso:
        print("Revalidar Acesso!")
    elif conceder_novo_acesso:
        print("Conceder novo Acesso!")

    

    if revalidar_acesso:
        sucesso, mensagem = revalida_acesso(driver, validade, worksheet, linha, workbook, file_path)
        if sucesso:
            print("Acesso revalidado com sucesso!")
            linha += 1
            continue
        else:
            print(f"Falha ao revalidar acesso. Mensagem: {mensagem}")
            linha += 1
            continue
    elif conceder_novo_acesso:
        sucesso, mensagem = atribuir_novo_acesso(driver, validade)
        if sucesso:
            print("Acesso concedido com sucesso!")
            linha += 1
            continue
        else:
            print(f"Falha ao conceder acesso. Mensagem: {mensagem}")
            linha += 1
            continue
    else:
        worksheet.cell(row=linha, column=7).value = f"Sistema já revalidado ({validade.strftime('%d/%m/%Y')})"
        print(f"Sistema já revalidado ({validade.strftime('%d/%m/%Y')})")
        linha += 1        
        try:
            workbook.save(file_path)
        except Exception as e:
            print(f"Erro ao salvar o arquivo: {e}")
        continue
