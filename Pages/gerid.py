import queue
import subprocess
import threading
import time
import tkinter as tk
from datetime import datetime
from tkinter import messagebox

import openpyxl
import pyotp
from selenium import webdriver
from selenium.common.exceptions import (ElementNotInteractableException,
                                        JavascriptException,
                                        NoSuchElementException, StaleElementReferenceException,
                                        TimeoutException)
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
import statistics
from datetime import datetime, timedelta
import os
import json
import numpy as np
from collections import deque

class GeridProcessoManager:
    def __init__(self, janela_movel=10):
        self.tempos_processamento = []
        self.tempos_movel = deque(maxlen=janela_movel)  # Últimos N tempos para média móvel
        self.total_registros = 0
        self.registros_processados = 0
        self.tempo_inicio_total = None
        self.janela_movel = janela_movel
        self.worksheet = None  # Adicionado para armazenar a referência à planilha

    def inicializar_contagem(self, worksheet):
        """Conta quantos registros precisam ser processados."""
        self.total_registros = 0
        linha = 2
        self.worksheet = worksheet  # Armazena referência à planilha para usar depois

        while worksheet.cell(row=linha, column=1).value is not None:
            if worksheet.cell(row=linha, column=8).value is None:  # Coluna onde status é salvo
                self.total_registros += 1
            linha += 1
        self.tempo_inicio_total = datetime.now()

    def atualizar_total_registros(self, linha_atual):
        """Atualiza o número de registros restantes baseado na linha atual."""
        if self.worksheet is None:
            return

        registros_restantes = 0
        linha_temp = linha_atual

        while self.worksheet.cell(row=linha_temp, column=1).value is not None:
            if self.worksheet.cell(row=linha_temp, column=8).value is None:  # Coluna onde status é salvo
                registros_restantes += 1
            linha_temp += 1

        return registros_restantes

    def registrar_tempo(self, tempo_inicio, tempo_fim, linha_atual=None):
        """Registra o tempo de processamento de um registro."""
        tempo_processamento = (tempo_fim - tempo_inicio).total_seconds()
        self.tempos_processamento.append(tempo_processamento)
        self.tempos_movel.append(tempo_processamento)

        # Atualiza registros processados usando a linha atual, se fornecida
        if linha_atual and self.worksheet:
            registros_restantes = self.atualizar_total_registros(linha_atual)
            self.registros_processados = self.total_registros - registros_restantes
        else:
            self.registros_processados += 1

    def calcular_tendencia(self):
        """Calcula se o processamento está acelerando ou desacelerando."""
        if len(self.tempos_processamento) < 5:
            return None, 0
        # Comparar a média dos últimos registros com a média geral
        media_geral = statistics.mean(self.tempos_processamento)
        media_recente = statistics.mean(self.tempos_movel)
        diferenca_percentual = ((media_geral - media_recente) / media_geral) * 100
        if diferenca_percentual > 5:
            return "acelerando", diferenca_percentual
        elif diferenca_percentual < -5:
            return "desacelerando", diferenca_percentual
        else:
            return "estável", diferenca_percentual

    def calcular_tempo_estimado(self, linha_atual=None):
        """Calcula tempo estimado para finalizar o processamento com base nos tempos registrados."""
        if not self.tempos_processamento:
            return "Calculando...", None, None, None

        # Atualizar contagem de registros restantes se uma linha atual for fornecida
        if linha_atual and self.worksheet:
            registros_restantes = self.atualizar_total_registros(linha_atual)
        else:
            registros_restantes = self.total_registros - self.registros_processados

        # Usar a média móvel se tiver registros suficientes, senão usar a média geral
        if len(self.tempos_movel) >= min(3, self.janela_movel):
            tempo_medio = statistics.mean(self.tempos_movel)
        else:
            tempo_medio = statistics.mean(self.tempos_processamento)

        # Calcular estimativa com base na tendência
        tendencia, percentual = self.calcular_tendencia() if len(self.tempos_processamento) >= 5 else (None, 0)

        # Ajustar a estimativa de acordo com a tendência
        fator_ajuste = 1.0
        if tendencia == "acelerando":
            fator_ajuste = max(0.8, 1.0 - (abs(percentual) / 200))
        elif tendencia == "desacelerando":
            fator_ajuste = min(1.2, 1.0 + (abs(percentual) / 200))

        tempo_medio_ajustado = tempo_medio * fator_ajuste
        tempo_total_estimado = tempo_medio_ajustado * registros_restantes

        # Calcular horário previsto de conclusão
        agora = datetime.now()
        horario_conclusao = agora + timedelta(seconds=tempo_total_estimado)

        # Formatar tempo restante
        horas = int(tempo_total_estimado // 3600)
        minutos = int((tempo_total_estimado % 3600) // 60)
        segundos = int(tempo_total_estimado % 60)

        # Calcular progresso
        progresso = ((self.total_registros - registros_restantes) / self.total_registros * 100) if self.total_registros > 0 else 0

        # Formatar mensagem com base no tempo estimado
        if horas > 0:
            mensagem = f"Estimativa: {horas}h {minutos}min {segundos}s"
        elif minutos > 0:
            mensagem = f"Estimativa: {minutos}min {segundos}s"
        else:
            mensagem = f"Estimativa: {segundos}s"

        # Adicionar info de horário de conclusão
        horario_str = horario_conclusao.strftime("%H:%M:%S")
        data_str = horario_conclusao.strftime("%d/%m/%Y")
        hoje_str = agora.strftime("%d/%m/%Y")

        if data_str == hoje_str:
            mensagem += f" (Término às {horario_str})"
        else:
            mensagem += f" (Término em {data_str} às {horario_str})"

        return mensagem, tendencia, progresso, horario_conclusao

    def calcular_velocidade_media(self):
        """Calcula a velocidade média de processamento (registros por minuto)."""
        if not self.tempos_processamento:
            return 0
        tempo_medio_segundos = statistics.mean(self.tempos_processamento)
        if tempo_medio_segundos > 0:
            return 60 / tempo_medio_segundos  # registros por minuto
        return 0

    def calcular_tempo_decorrido(self):
        """Calcula o tempo decorrido desde o início do processamento."""
        if self.tempo_inicio_total:
            decorrido = datetime.now() - self.tempo_inicio_total
            horas = int(decorrido.total_seconds() // 3600)
            minutos = int((decorrido.total_seconds() % 3600) // 60)
            segundos = int(decorrido.total_seconds() % 60)
            if horas > 0:
                return f"{horas}h {minutos}min {segundos}s"
            elif minutos > 0:
                return f"{minutos}min {segundos}s"
            else:
                return f"{segundos}s"
        return "Desconhecido"

    def get_status(self, linha_atual=None):
        """Retorna o status atual do processamento com informações detalhadas."""
        tempo_estimado, tendencia, progresso, _ = self.calcular_tempo_estimado(linha_atual)

        if linha_atual and self.worksheet:
            registros_restantes = self.atualizar_total_registros(linha_atual)
            registros_processados = self.total_registros - registros_restantes
        else:
            registros_restantes = self.total_registros - self.registros_processados
            registros_processados = self.registros_processados

        velocidade = self.calcular_velocidade_media()
        tempo_decorrido = self.calcular_tempo_decorrido()

        # Construir mensagem de status
        status = f"{progresso:.1f}% ({registros_processados}/{self.total_registros}) | "
        status += f"{tempo_estimado} | Restantes: {registros_restantes} | "
        status += f"{velocidade:.1f} reg/min | Tempo decorrido: {tempo_decorrido}"

        # Adicionar informação de tendência se disponível
        if tendencia:
            status += f" | Processamento {tendencia}"
        return status

def safe_find_element(driver, by, value, max_attempts=5, delay=2):
    """Função para localizar um elemento de forma segura."""
    attempts = 0
    while attempts < max_attempts:
        try:
            return driver.find_element(by, value)
        except StaleElementReferenceException:
            print(f"StaleElementReferenceException ao tentar encontrar elemento {value}. Tentativa {attempts + 1} de {max_attempts}.")
            attempts += 1
            time.sleep(delay)
        except NoSuchElementException:
            print(f"Elemento {value} não encontrado.")
            break
    raise Exception(f"Não foi possível encontrar o elemento {value} após {max_attempts} tentativas.")

def show_success_popup():
    root = tk.Tk()
    root.withdraw()  # Esconde a janela principal
    messagebox.showinfo("Sucesso", "A operação foi executada com sucesso!")
    root.destroy()

def close_workbook(workbook):
    try:
        workbook.close()
        print("Arquivo Excel fechado com sucesso.")
    except Exception as e:
        print(f"Erro ao fechar o arquivo Excel: {e}")

def execute_javascript_with_retry(driver, script, max_attempts=3, refresh_delay=10, retry_delay=5):
    attempts = 0
    while attempts < max_attempts:
        try:
            driver.execute_script(script)
            #print("Script JavaScript executado com sucesso.")
            return True
        except JavascriptException as js_error:
            print(f"Erro ao executar script JavaScript: {js_error}. Tentando novamente em {retry_delay} segundos...")
            time.sleep(retry_delay)
        except NoSuchElementException:
            print("Elemento não encontrado, mas continuando a execução.")
            return False
        except Exception as e:
            print(f"Ocorreu um erro: {e}. Atualizando a página e repetindo operação em {refresh_delay} segundos.")
            driver.refresh()
            time.sleep(refresh_delay)

        attempts += 1

    print(f"Falha ao executar o script após {max_attempts} tentativas.")
    return False

def verificar_mensagem_operacao(driver, worksheet, linha):
    mensagens = {
        "/html/body/div[1]/div[2]/form[1]/div[2]/ul/li": [
            "A operação foi executada com sucesso."
        ],
        "/html/body/div[1]/div[2]/ul/li": [
            "Domínio não existe.",
            "Ocorreu um erro de comunicação. Aguarde alguns minutos e tente novamente.",
            "A Data de Validade não deve ser superior a Data de Validade da credencial do usuário emissor.",
            "Gestor de Acesso só pode atribuir acesso no seu próprio domínio ou domínio abaixo de sua abrangência.",
            "Não é permitido dar uma autorização a si mesmo."
        ]
    }

    for xpath, textos_esperados in mensagens.items():
        try:
            elements = driver.find_elements(By.XPATH, xpath)
            if len(elements) == 1 and elements[0].text in textos_esperados:
                mensagem = elements[0].text
                print(mensagem)
                worksheet.cell(row=linha, column=8).value = mensagem
                return True
        except:
            continue

    return False

def element_exists(driver, xpath, max_attempts=2, delay=1):
    """Verifica se um elemento existe na página usando JavaScript, com retentativas."""
    js_script = """
        return document.evaluate(
            arguments[0],
            document,
            null,
            XPathResult.FIRST_ORDERED_NODE_TYPE,
            null
        ).singleNodeValue !== null;
    """
    attempts = 0
    while attempts < max_attempts:
        if driver.execute_script(js_script, xpath):
            return True
        attempts += 1
        time.sleep(delay)
    return False

def check_element_with_retry(driver, xpath, max_attempts=3, delay=1):
    """Verifica a presença de um elemento usando JavaScript, com retentativas."""
    js_script = """
        return document.evaluate(
            arguments[0],
            document,
            null,
            XPathResult.FIRST_ORDERED_NODE_TYPE,
            null
        ).singleNodeValue;
    """
    attempts = 0
    while attempts < max_attempts:
        try:
            element = driver.execute_script(js_script, xpath)
            if element:
                # Converter o elemento retornado pelo JavaScript em um WebElement do Selenium
                web_element = driver.find_element(By.XPATH, xpath)
                return web_element.text
        except Exception as e:
            print(f"Erro ao executar script JavaScript: {e}. Tentando novamente em {delay} segundo(s)...")
        attempts += 1
        time.sleep(delay)
    return None

def run_automation_gerid(file_path, update_label_func=None, update_status_func=None, stop_event=None):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    processo_manager = GeridProcessoManager()
    processo_manager.inicializar_contagem(worksheet)
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    Finaliza = True
    
    try:
        linha = 2
        coluna = 3
        driver.get("https://geridinss.dataprev.gov.br/gpa")
        driver.implicitly_wait(10)
        driver.maximize_window()
        print(f"Aguardando procedimento de Login...")
        WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/ul/li")))
        print(f"Login realizado com sucesso.")
        success = execute_javascript_with_retry(driver, "document.getElementById('formMenu:btAtribAcesso').click();")
        if success:
            print("O botão foi clicado com sucesso.")
        else:
            print("Não foi possível clicar no botão após várias tentativas.")
        if update_status_func:
            update_status_func("Processando registros...")
        
        while Finaliza and not stop_event.is_set():
            try:
                if stop_event and stop_event.is_set():
                    print("Interrupção solicitada, finalizando...")
                    break
                servidor = worksheet.cell(row=linha, column=1).value
                UO = worksheet.cell(row=linha, column=2).value
                Sistema = worksheet.cell(row=linha, column=coluna).value
                Subsistema = worksheet.cell(row=linha, column=coluna + 1).value
                Papel = worksheet.cell(row=linha, column=coluna + 2).value
                TipoUO = worksheet.cell(row=linha, column=coluna + 3).value
                validade = worksheet.cell(row=linha, column=coluna + 4).value
                if validade:
                    validade_str = validade.strftime("%d/%m/%Y")
                Situacao = worksheet.cell(row=linha, column=coluna + 5).value
                if not servidor:
                    try:
                        workbook.save(file_path)
                        print("Arquivo salvo com sucesso.")
                        close_workbook(workbook)
                    except Exception as e:
                        print(f"Erro ao salvar ou fechar o arquivo: {e}")
                    finally:
                        driver.quit()
                        print("Final!")
                        show_success_popup()
                        return
                

                if Situacao is not None:
                    linha += 1
                    if update_label_func:
                        update_label_func(linha)
                    continue



                if update_label_func:
                    update_label_func(linha)



                print("==============================================================")
                print(f"{linha}:{servidor}-{UO}-{Sistema}-{Subsistema}-{Papel}-{validade}")
                driver.get("https://geridinss.dataprev.gov.br/gpa")
                tempo_inicio = datetime.now()
                time.sleep(2)
                success = execute_javascript_with_retry(driver, "document.getElementById('formMenu:btAtribAcesso').click();")
                if success:
                    pass
                else:
                    print("Não foi possível clicar no botão após várias tentativas.")
                    time.sleep(5)
                    continue
                sistema_label_xpath = "/html/body/div[1]/div[2]/form[1]/fieldset/div[1]/label/span"
                try:
                    sistema_label = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, sistema_label_xpath)))
                    try:
                        Select(safe_find_element(driver, By.ID, "form:sistema")).select_by_visible_text(Sistema)
                    except (NoSuchElementException, ElementNotInteractableException):
                        print("Sistema não localizado dentre as opções disponíveis.")
                        worksheet.cell(row=linha, column=8).value = "Sistema não localizado dentre as opções disponíveis."
                        linha += 1
                        continue
                except TimeoutException:
                    driver.get("https://geridinss.dataprev.gov.br/gpa")
                    success = execute_javascript_with_retry(driver, "document.getElementById('formMenu:btAtribAcesso').click();")
                    if success:
                        pass
                    else:
                        print("Não foi possível clicar no botão após várias tentativas.")
                    continue
                try:
                    Select(safe_find_element(driver, By.ID, "form:subsistema")).select_by_visible_text(Subsistema)
                    Select(safe_find_element(driver, By.ID, "form:papel")).select_by_visible_text(Papel)
                    Select(safe_find_element(driver, By.ID, "form:tipoDominio")).select_by_visible_text(TipoUO)
                    driver.find_element(By.ID, "form:dominio").clear()
                    driver.find_element(By.ID, "form:dominio").send_keys(UO)
                    driver.find_element(By.ID, "form:usuario").clear()
                    driver.find_element(By.ID, "form:usuario").send_keys(servidor)
                    driver.find_element(By.ID, "form:filtrar").click()
                    time.sleep(1)
                except (NoSuchElementException, ElementNotInteractableException) as e:
                    print(f"Erro ao interagir com elementos da página: {e}")
                    print("Atualizando a página e tentando novamente...")
                    driver.get("https://geridinss.dataprev.gov.br/gpa")
                    time.sleep(5)
                    continue
                xpath = "/html/body/div[1]/div[2]/ul/li"
                js_script = """
                    return document.evaluate(
                        arguments[0],
                        document,
                        null,
                        XPathResult.FIRST_ORDERED_NODE_TYPE,
                        null
                    ).singleNodeValue !== null;
                """
                try:
                    elemento_existe = driver.execute_script(js_script, xpath)
                    if elemento_existe:
                        print("Erro de comunicação detectado. Aguardando e tentando novamente...")
                        driver.get("https://geridinss.dataprev.gov.br/gpa")
                        execute_javascript_with_retry(driver, "document.getElementById('formMenu:btAtribAcesso').click();")
                        continue
                except:
                    pass
                try:


                    if element_exists(driver, "/html/body/div[1]/div[2]/form[2]/table/tbody/tr/td[7]"):
                        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/form[2]/table/tbody/tr/td[7]")))
                        data_string = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/form[2]/table/tbody/tr/td[7]").text
                        data = datetime.strptime(data_string, "%d/%m/%Y")
                        validade_formatada = validade.strftime("%d/%m/%Y")
                        data_formatada = data.strftime("%d/%m/%Y")
                        print("Data de validade informada:", validade_formatada)
                        print("Data de validade atual:", data_formatada)
                        data_string_dt = datetime.strptime(data_string, "%d/%m/%Y")
                        try:
                            workbook.save(file_path)
                        except Exception as e:
                            print(f"Erro ao salvar o arquivo: {e}")


                        # if data_string_dt < validade:
                        #     print("A data extraída do gerid é menor que a data informada. Avançando para revalidação...")
                        #     driver.find_element(By.ID, "dataTableCredencial:selected").click()
                        #     driver.find_element(By.ID, "form2:btAlterar").click()
                        #     driver.find_element(By.ID, "form2:dataValidade").send_keys(Keys.CONTROL, 'a')
                        #     validade_str = validade.strftime("%d/%m/%Y")
                        #     driver.find_element(By.ID, "form2:dataValidade").send_keys(validade_str)
                        #     driver.find_element(By.ID, "form2:confirmar").click()
                        #     sucesso, nova_coluna = verificar_mensagem_operacao(driver, worksheet, linha, coluna)
                        #     if sucesso:
                        #         coluna = nova_coluna
                        #         continue
                        #     tempo_fim = datetime.now()
                        #     processo_manager.registrar_tempo(tempo_inicio, tempo_fim, linha)
                        #     if update_status_func:
                        #         update_status_func(processo_manager.get_status(linha))
                        #     try:
                        #         workbook.save(file_path)
                        #     except Exception as e:
                        #         print(f"Erro ao salvar o arquivo: {e}")
                        # else:
                        #     worksheet.cell(row=linha, column=coluna + 5).value = f"Sistema já revalidado ({validade.strftime('%d/%m/%Y')})"
                        #     print(f"Sistema já revalidado ({validade.strftime('%d/%m/%Y')})")
                        #     tempo_fim = datetime.now()
                        #     processo_manager.registrar_tempo(tempo_inicio, tempo_fim, linha)
                        #     if update_status_func:
                        #         update_status_func(processo_manager.get_status(linha))
                        #     try:
                        #         workbook.save(file_path)
                        #     except Exception as e:
                        #         print(f"Erro ao salvar o arquivo: {e}")
                    
                    
                        #print("A data extraída do gerid é menor que a data informada. Avançando para revalidação...")
                        
                        if data_string_dt < validade:
                            driver.find_element(By.ID, "dataTableCredencial:selected").click()
                            driver.find_element(By.ID, "form2:btAlterar").click()
                            driver.find_element(By.ID, "form2:dataValidade").send_keys(Keys.CONTROL, 'a')
                            validade_str = validade.strftime("%d/%m/%Y")
                            driver.find_element(By.ID, "form2:dataValidade").send_keys(validade_str)
                            driver.find_element(By.ID, "form2:confirmar").click()
                            sucesso = verificar_mensagem_operacao(driver, worksheet, linha)
                            if sucesso:
                                linha += 1               
                                pass
                            else:
                                print("Erro - Repetindo operação!")
                                continue
                            tempo_fim = datetime.now()

                            processo_manager.registrar_tempo(tempo_inicio, tempo_fim, linha)
                            if update_status_func:
                                update_status_func(processo_manager.get_status(linha))
                            try:
                                workbook.save(file_path)
                            except Exception as e:
                                print(f"Erro ao salvar o arquivo: {e}")
                        else:
                            worksheet.cell(row=linha, column=coluna + 5).value = f"Sistema já revalidado ({validade.strftime('%d/%m/%Y')})"
                            print(f"Sistema já revalidado ({validade.strftime('%d/%m/%Y')})")
                            tempo_fim = datetime.now()
                            processo_manager.registrar_tempo(tempo_inicio, tempo_fim, linha)
                            driver.get("https://geridinss.dataprev.gov.br/gpa")
                            time.sleep(3)
                            success = execute_javascript_with_retry(driver, "document.getElementById('formMenu:btAtribAcesso').click();")
                            linha += 1
                            continue


                    
                    else:
                        def retry_on_stale_element(func, *args, max_attempts=5, delay=1):
                            """Tenta executar uma função e repete em caso de StaleElementReferenceException."""
                            attempts = 0
                            while attempts < max_attempts:
                                try:
                                    return func(*args)
                                except StaleElementReferenceException:
                                    attempts += 1
                                    print(f"StaleElementReferenceException ocorreu. Tentativa {attempts} de {max_attempts}.")
                                    time.sleep(delay)
                            raise Exception(f"Não foi possível executar a ação após {max_attempts} tentativas.")

                        # Exemplo de uso da função retry_on_stale_element
                        def safe_interaction(driver, by, value, action):
                            """Interage com um elemento de forma segura, repetindo em caso de StaleElementReferenceException."""
                            def interaction():
                                element = driver.find_element(by, value)
                                action(element)
                            retry_on_stale_element(interaction)

                        # Aplicando a lógica no trecho do código
                        if element_exists(driver, "/html/body/div[1]/div[2]/form[1]/div[2]/ul/li"):
                            print("Atribuindo novo acesso....")
                            try:
                                # Clicar no botão "Novo"
                                retry_on_stale_element(lambda: driver.find_element(By.ID, "form2:novo").click())
                                time.sleep(2)
                                # Selecionar Sistema
                                retry_on_stale_element(lambda: Select(safe_find_element(driver, By.ID, "form:sistema")).select_by_visible_text(Sistema))

                                # Selecionar Subsistema
                                retry_on_stale_element(lambda: Select(safe_find_element(driver, By.ID, "form:subsistema")).select_by_visible_text(Subsistema))

                                # Selecionar Papel
                                retry_on_stale_element(lambda: Select(safe_find_element(driver, By.ID, "form2:papel")).select_by_visible_text(Papel))

                                # Selecionar Tipo de Domínio
                                retry_on_stale_element(lambda: Select(safe_find_element(driver, By.ID, "form2:tipoDominio")).select_by_visible_text(TipoUO))

                                # Preencher Domínio
                                def fill_dominio(element):
                                    element.clear()
                                    element.send_keys(UO)
                                safe_interaction(driver, By.ID, "form2:dominio", fill_dominio)

                                # Preencher Usuário
                                def fill_usuario(element):
                                    element.click()
                                    element.clear()
                                    element.send_keys(servidor)
                                safe_interaction(driver, By.ID, "form2:usuario", fill_usuario)

                                # Preencher Data de Validade
                                def fill_data_validade(element):
                                    element.click()
                                    element.clear()
                                    element.send_keys(Keys.CONTROL, 'a')
                                    validade_str = validade.strftime("%d/%m/%Y")
                                    element.send_keys(validade_str)
                                safe_interaction(driver, By.ID, "form2:dataValidade", fill_data_validade)

                                # Selecionar Períodos
                                for periodo in ["form2:periodo:0", "form2:periodo:6", "form2:periodo:7"]:
                                    retry_on_stale_element(lambda: safe_find_element(driver, By.ID, periodo).click())

                                # Preencher Hora de Início
                                def fill_hora_inicio(element):
                                    element.click()
                                    element.send_keys(Keys.CONTROL, 'a')
                                    element.send_keys("0000")
                                safe_interaction(driver, By.ID, "form2:horaAcessoInicio", fill_hora_inicio)

                                # Preencher Hora de Fim
                                def fill_hora_fim(element):
                                    element.click()
                                    element.send_keys(Keys.CONTROL, 'a')
                                    element.send_keys("2359")
                                safe_interaction(driver, By.ID, "form2:horaAcessoFim", fill_hora_fim)

                                # Clicar em Confirmar
                                retry_on_stale_element(lambda: safe_find_element(driver, By.ID, "form2:confirmar").click())

                                # Verificar mensagem de sucesso
                                mensagem = check_element_with_retry(driver, "/html/body/div[1]/div[2]/ul/li")
                                if mensagem:
                                    print(mensagem)
                                    worksheet.cell(row=linha, column=8).value = mensagem

                                # Verificar mensagem de sucesso novamente
                                mensagem_sucesso = check_element_with_retry(driver, "/html/body/div[1]/div[2]/form[1]/div[2]/ul/li")
                                if mensagem_sucesso:
                                    print(mensagem_sucesso)
                                    worksheet.cell(row=linha, column=8).value = mensagem_sucesso

                                print(f"Acesso: {UO} - {Sistema} - {Subsistema} - {Papel} - {servidor} - {validade_str}")

                                try:
                                    workbook.save(file_path)
                                except Exception as e:
                                    print(f"Erro ao salvar o arquivo: {e}")

                                tempo_fim = datetime.now()
                                processo_manager.registrar_tempo(tempo_inicio, tempo_fim, linha)
                                if update_status_func:
                                    update_status_func(processo_manager.get_status(linha))
                                linha += 1
                                continue
                            except Exception as e:
                                print(f"Erro ao atribuir novo acesso repetindo operação: {e}")
                                driver.get("https://geridinss.dataprev.gov.br/gpa")
                                time.sleep(3)
                                success = execute_javascript_with_retry(driver, "document.getElementById('formMenu:btAtribAcesso').click();")
                                if success:
                                    print("O botão foi clicado com sucesso.")
                                else:
                                    print("Não foi possível clicar no botão após várias tentativas.")
                                continue



                except Exception as e:
                    print(f"Erro inesperado: {e}")
                    print("Tentando recuperar e continuar...")
                    driver.get("https://geridinss.dataprev.gov.br/gpa")
                    time.sleep(3)
                    success = execute_javascript_with_retry(driver, "document.getElementById('formMenu:btAtribAcesso').click();")
                    if success:
                        print("O botão foi clicado com sucesso.")
                    else:
                        print("Não foi possível clicar no botão após várias tentativas.")
                    time.sleep(5)
                    continue
            except Exception as e:
                if stop_event.is_set():
                    break
                tempo_fim = datetime.now()
                processo_manager.registrar_tempo(tempo_inicio, tempo_fim, linha)
                if update_status_func:
                    update_status_func(processo_manager.get_status(linha))
                print(f"Erro inesperado: {e}")
                print("Tentando recuperar e continuar...")
                driver.get("https://geridinss.dataprev.gov.br/gpa")
                time.sleep(3)
                success = execute_javascript_with_retry(driver, "document.getElementById('formMenu:btAtribAcesso').click();")
                if success:
                    print("O botão foi clicado com sucesso.")
                else:
                    print("Não foi possível clicar no botão após várias tentativas.")
                time.sleep(5)
                continue
        print("==== FIM ====")
    except Exception as e:
        print(f"Erro geral: {e}")
        if update_status_func:
            update_status_func("Erro durante a execução...")
    finally:
        driver.quit()
        print("Driver encerrado com sucesso.")
        if stop_event.is_set():
            if update_status_func:
                update_status_func("Operação interrompida pelo usuário")
    if update_status_func:
        update_status_func("Execução finalizada.")

def run_automation_thread(file_path, update_label_func=None, update_status_func=None, stop_event=None):
    thread = threading.Thread(target=run_automation_gerid, args=(file_path, update_label_func, update_status_func, stop_event))
    thread.start()
    return thread
