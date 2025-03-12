from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, ElementNotInteractableException
from selenium.common.exceptions import JavascriptException, NoSuchElementException
from datetime import datetime
import openpyxl
import time
import tkinter as tk
from tkinter import messagebox
import subprocess
import threading
import queue
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
import time
import pyotp
from selenium.webdriver.support.ui import Select


class ReiniciarProcessoException(Exception):
    """Exceção personalizada para sinalizar a reinicialização do processo."""
    pass



def show_success_popup():
    root = tk.Tk()
    root.withdraw() # Esconde a janela principal
    messagebox.showinfo("Sucesso", "A operação foi executada com sucesso!")
    root.destroy()

def close_workbook(workbook):
    try:
        workbook.close()
        print("Arquivo Excel fechado com sucesso.")
    except Exception as e:
        print(f"Erro ao fechar o arquivo Excel: {e}")



def get_codigo_sv(worksheet, linha):
    CodigoSv = []
    coluna = 11
    while True:
        valor = worksheet.cell(row=linha, column=coluna).value
        if valor is None:
            break
        CodigoSv.append(valor)
        coluna += 1
    return CodigoSv

def print_values(siape, Unidade, BloquerAlteracoes, ResetarTodosSv, AreaMeio, GrupoMeio, Status, CodigoSv):
    print("============================================================")
    print(f"Matrícula:{siape}-{Unidade}-{Status}-Área Meio:{AreaMeio}")
    print("Bloquear Alterações para Unidades Inferiores:", BloquerAlteracoes)
    print("Resetar Todas Competências:", ResetarTodosSv)
    print("Código SV:", CodigoSv)
    print("configurando...")

def click_checkbox(driver, checkbox_id, condition):
    max_attempts = 2
    for attempt in range(max_attempts):
        try:
            checkbox = driver.find_element(By.ID, checkbox_id)
            if checkbox.is_selected() != condition:
                checkbox.click()
                return True
            return False
        except StaleElementReferenceException:
            if attempt < max_attempts - 1:
                print(f"Stale element ao clicar no checkbox {checkbox_id}. Tentando novamente...")
                time.sleep(2)  # Espera antes de recarregar o elemento
                continue
            else:
                print(f"Falha após {max_attempts} tentativas no checkbox {checkbox_id}")
                return False
        except Exception as e:
            print(f"Erro inesperado ao clicar no checkbox {checkbox_id}: {str(e)}")
            return False
    return False

def get_total_pages(driver):
    try:
        paginator_text = driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/span[4]/fieldset/div[2]/div[3]/div/div/div/div[2]/span[1]").text
        total_pages = int(paginator_text.split(" de ")[1])
        return total_pages
    except NoSuchElementException:
        return 1

def read_all_units(driver):
    wait = WebDriverWait(driver, 5)
    total_pages = get_total_pages(driver)
    unidades = []
    for page in range(1, total_pages + 1):
        contador = 1
        while True:
            try:
                xpath = f"/html/body/div[3]/div/form[1]/span[4]/fieldset/div[2]/div[3]/div/div/div/div/table/tbody/tr[{contador}]/td[2]"
                element = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
                NomeUnidade = element.text
                codigo = NomeUnidade.split('-')[0].strip()
                unidades.append(codigo)
                contador += 1
            except (TimeoutException, StaleElementReferenceException):
                break
        if page < total_pages:
            try:
                next_button = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "ui-paginator-next")))
                driver.execute_script("arguments[0].scrollIntoView(true);", next_button)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", next_button)
                wait.until(EC.staleness_of(element))
                time.sleep(1)
            except Exception as e:
                print(f"Erro ao mudar para página {page + 1}: {str(e)}")
                break
    return unidades

def wait_for_table_update(driver, cod, timeout=5):
    try:
        # Esperar até que a tabela seja atualizada com o novo código
        WebDriverWait(driver, timeout).until(
            EC.text_to_be_present_in_element(
                (By.XPATH, "/html/body/div[3]/div/form[1]/span[5]/fieldset/div[2]/div/div[2]/span/div/div[1]/table/tbody/tr/td[2]"),
                str(cod)
            )
        )
        print(f"Tabela atualizada com o código {cod}")
        return True
    except TimeoutException:
        print(f"Timeout: A tabela não foi atualizada com o código {cod} dentro do tempo esperado")
        return False

def find_and_process_unit(driver, unidade, max_retries=3):
    for attempt in range(max_retries):
        try:
            wait = WebDriverWait(driver, 10)
            try:
                wait = WebDriverWait(driver, 10)
                target_element = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div/form[1]/span[4]/fieldset/div[2]/div[3]/div/div/div/div[2]")))
                select_element = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div/form[1]/span[4]/fieldset/div[2]/div[3]/div/div/div/div[2]/select")))
                select = Select(select_element)
                select.select_by_value("30")
            except TimeoutException:
                print("Não possui paginação > avançando...")
            table = wait.until(EC.presence_of_element_located((By.ID, "form:tabelaUnidades_data")))
            rows = table.find_elements(By.TAG_NAME, "tr")
            for i, row in enumerate(rows):
                codigo_celula = wait.until(EC.presence_of_element_located((By.XPATH, f"//tbody[@id='form:tabelaUnidades_data']/tr[{i+1}]/td[2]"))).text.split(" - ")[0]
                if codigo_celula == unidade:
                    print(f"Unidade {unidade} encontrada na linha {i+1}")
                    # Verificar e marcar o checkbox GET se necessário
                    checkbox_id = f"form:tabelaUnidades:{i}:selecionarDeselecionarGet"
                    checkbox = wait.until(EC.presence_of_element_located((By.ID, checkbox_id)))
                    if not checkbox.is_selected():
                        driver.execute_script("arguments[0].click();", checkbox)
                        print(f"Checkbox GET marcado para a unidade {unidade}")
                        time.sleep(1)
                    # Clicar no botão de edição
                    edit_button_xpath = f"//a[contains(@id, 'form:tabelaUnidades:{i}:j_idt') and @class='ui-commandlink ui-widget btn ico-pencil']"
                    edit_button = wait.until(EC.element_to_be_clickable((By.XPATH, edit_button_xpath)))
                    driver.execute_script("arguments[0].click();", edit_button)
                    print(f"Botão de edição clicado para a unidade {unidade}")
                    return True
            print(f"Unidade {unidade} não encontrada na tabela")
            return False
        except (StaleElementReferenceException, TimeoutException) as e:
            if attempt < max_retries - 1:
                print(f"Tentativa {attempt + 1} falhou. Tentando novamente...")
                time.sleep(2)
            else:
                print(f"Erro após {max_retries} tentativas: {str(e)}")
                return False

def verificar_e_bloquear_alteracoes(driver, bloquear_alteracoes):
    try:
        wait = WebDriverWait(driver, 10)
        checkbox_nao_bloqueado = wait.until(EC.presence_of_element_located((By.ID, "cmpModalCompetenciaServicoLocal:formPesquisaCompetencias:bloquearAlteracaoExercicio:0")))
        esta_bloqueado = not checkbox_nao_bloqueado.is_selected()
        print(f"Estado atual: {'Bloqueado' if esta_bloqueado else 'Não bloqueado'}")
        if bloquear_alteracoes == "Sim" and not esta_bloqueado:
            checkbox_bloqueado = driver.find_element(By.ID, "cmpModalCompetenciaServicoLocal:formPesquisaCompetencias:bloquearAlteracaoExercicio:1")
            driver.execute_script("arguments[0].click();", checkbox_bloqueado)
            print("Alterações bloqueadas.")
        elif bloquear_alteracoes == "Não" and esta_bloqueado:
            checkbox_nao_bloqueado = driver.find_element(By.ID, "cmpModalCompetenciaServicoLocal:formPesquisaCompetencias:bloquearAlteracaoExercicio:0")
            driver.execute_script("arguments[0].click();", checkbox_nao_bloqueado)
            print("Alterações desbloqueadas.")
        elif bloquear_alteracoes == "Sim" and esta_bloqueado:
            print("Já está bloqueado. Nenhuma ação necessária.")
        else:
            print("Não está bloqueado e não precisa ser bloqueado. Nenhuma ação necessária.")
        time.sleep(1)
    except Exception as e:
        print(f"Erro ao verificar ou alterar o bloqueio: {str(e)}")

def retry_find_element(driver, xpath, max_attempts=20):
    for attempt in range(max_attempts):
        print(f"{attempt} - Erro stale element efetuando nova tentativa de localização..")
        try:
            wait = WebDriverWait(driver, 10)
            element = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
            return element
        except StaleElementReferenceException:
            if attempt == max_attempts - 1:
                raise
            time.sleep(1)

def process_code(driver, code, ativar_mi_exer, input_xpath):
    max_attempts = 20
    for attempt in range(max_attempts):
        try:
            wait = WebDriverWait(driver, 10)
            input_field = wait.until(EC.presence_of_element_located((By.XPATH, input_xpath)))
            input_field.clear()
            time.sleep(0.5)
            input_field.send_keys(str(code))
            print(f"Código {code} inserido com sucesso.")
            time.sleep(1)
            # Aguardar a tabela aparecer
            checkbox_id = "cmpModalCompetenciaServicoLocal:formPesquisaCompetencias:tabelaServicoModal:0:selecionarDeselecionarCompetencia"
            checkbox = wait.until(EC.presence_of_element_located((By.ID, checkbox_id)))
            is_checked = checkbox.is_selected()
            if ativar_mi_exer == "Sim":
                if not is_checked:
                    checkbox.click()
                    print(f"Checkbox marcado para o código {code}")
            elif ativar_mi_exer == "Não":
                if is_checked:
                    checkbox.click()
                    print(f"Checkbox desmarcado para o código {code}")
            else:
                print(f"Nenhuma ação necessária para o código {code}")
            time.sleep(1)
            break # Sucesso, sair do loop
        except StaleElementReferenceException:
            print(f"Tentativa {attempt + 1}: Elemento stale, tentando novamente...")
            time.sleep(1)
        except TimeoutException:
            print(f"Tentativa {attempt + 1}: Timeout ao tentar encontrar o elemento, tentando novamente...")
            time.sleep(1)
        except Exception as e:
            print(f"Tentativa {attempt + 1}: Ocorreu um erro: {e}")
            time.sleep(1)
        if attempt == max_attempts - 1:
            print(f"Não foi possível interagir com o elemento após {max_attempts} tentativas.")
            raise ReiniciarProcessoException("Reiniciar o processo principal")

def confirmar_e_processar(driver, worksheet, linha):
    try:
        # Clicar no botão confirmar principal
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        wait = WebDriverWait(driver, 10)
        botao_confirmar = wait.until(EC.element_to_be_clickable((By.ID, "form:botaoConfirmar")))
        botao_confirmar.click()
        print("Botão Confirmar principal clicado com sucesso.")
        time.sleep(2)
        # Verificar se aparece a mensagem de divergência
        mensagem_xpath = "/html/body/div[3]/div/div[2]/div/ul/li/span"
        try:
            mensagem = wait.until(EC.presence_of_element_located((By.XPATH, mensagem_xpath)))
            if "Horário diverge dos dados do servidor oriundo do SISREF" in mensagem.text:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                print("Mensagem de divergência detectada. Confirmando novamente.")
                botao_confirmar = wait.until(EC.element_to_be_clickable((By.ID, "form:botaoConfirmar")))
                botao_confirmar.click()
                print("Segundo clique no botão Confirmar realizado.")
                time.sleep(2)
        except:
            print("Não foi detectada mensagem de divergência.")
        # Verificar se a alteração foi realizada com sucesso
        try:
            mensagem_sucesso = wait.until(EC.presence_of_element_located((By.XPATH, mensagem_xpath)))
            if "Alteração realizada(o) com sucesso." in mensagem_sucesso.text:
                print("Alteração realizada com sucesso!")
                worksheet.cell(row=linha, column=10).value = "Alteração realizada com sucesso!"
                return True
            else:
                print(f"Mensagem inesperada: {mensagem_sucesso.text}")
                worksheet.cell(row=linha, column=10).value = f"Mensagem inesperada: {mensagem_sucesso.text}"
                return False
        except:
            print("Não foi possível encontrar mensagem de confirmação.")
            worksheet.cell(row=linha, column=10).value = "Não foi possível encontrar mensagem de confirmação."
            return False
    except Exception as e:
        print(f"Erro durante o processo de confirmação: {str(e)}")
        return False

def run_automation_saggestao(file_path, update_label_func=None, update_status_func=None):
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    update_status_func("Iniciando execução.")
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    driver.get("http://psagapr01/saggestaoagu/pages/cadastro/profissional/consultar.xhtml")
    driver.implicitly_wait(10)
    driver.maximize_window()
    
    #==== trecho exclusivo para RICARDO COMENTAR DEPOIS====================================
    print(f"Aguardando Tela Código OTP")
    WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.XPATH, "/html/body/div/div/main/div/div/div/div[2]/form/div/section[2]/input")))
    print(f"Login realizado com sucesso.")
    totp_key = "SCFW2ZRV7FZ4G3YO"
    totp = pyotp.TOTP(totp_key)
    codigo_totp = totp.now()
    print(f"Código TOTP: {codigo_totp}")
    wait = WebDriverWait(driver, 10)
    driver.find_element(By.XPATH, "/html/body/div/div/main/div/div/div/div[2]/form/div/section[2]/input").click()
    wait = WebDriverWait(driver, 10)
    driver.find_element(By.XPATH, "/html/body/div/div/main/div/div/div/div[2]/form/div/section[2]/input").send_keys(codigo_totp)
    wait = WebDriverWait(driver, 10)
    driver.find_element(By.XPATH, "/html/body/div/div/main/div/div/div/div[2]/form/div/section[3]/input").click()
    wait = WebDriverWait(driver, 10)
    driver.find_element(By.XPATH, "/html/body/div/div[2]/button[3]").click()
    wait = WebDriverWait(driver, 10)
    driver.find_element(By.XPATH, "/html/body/div/div[3]/p[2]/a").click()
    wait = WebDriverWait(driver, 10)
    select_element = wait.until(EC.presence_of_element_located((By.ID, "domains")))
    select = Select(select_element)
    select.select_by_value("UO:01.001.PRES")
    submit_button = driver.find_element(By.CSS_SELECTOR, "input[type='submit'][value='Enviar']")
    submit_button.click()
    #===finalizar minha parte=======================
    
    print(f"Aguardando Login")
    WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div/form[1]/div[1]/div[2]/div[1]/div[1]/div[1]/div/span[2]")))
    print(f"Login realizado com sucesso.")
    linha = 2 # Defina a linha inicial conforme necessário
    
    while True:
        update_label_func(linha)
        update_status_func("Em execução...") 
        skip_to_next_iteration = False  # Variável de controle       
        siape = worksheet.cell(row=linha, column=1).value
        Unidade = worksheet.cell(row=linha, column=2).value
        AtribResp = worksheet.cell(row=linha, column=3).value
        Trasf = worksheet.cell(row=linha, column=4).value
        AtivarMiExer = worksheet.cell(row=linha, column=5).value
        BloquerAlteracoes = worksheet.cell(row=linha, column=6).value
        ResetarTodosSv = worksheet.cell(row=linha, column=7).value
        AreaMeio = worksheet.cell(row=linha, column=8).value
        GrupoMeio = worksheet.cell(row=linha, column=9).value
        Status = worksheet.cell(row=linha, column=10).value
        CodigoSv = get_codigo_sv(worksheet, linha)
        
        if Status is not None:
            linha += 1
            continue
        
        # Imprimindo todos os valores
        print_values(siape, Unidade, BloquerAlteracoes, ResetarTodosSv, AreaMeio, GrupoMeio, Status, CodigoSv)
        
        if not siape:
            try:
                workbook.save(file_path)
                print("Arquivo salvo com sucesso.")
                close_workbook(workbook)
            except Exception as e:
                print(f"Erro ao salvar ou fechar o arquivo: {e}")
            finally:
                driver.quit()
                print("Final!")
                show_success_popup()
                return
        
        
        if not siape:
            try:
                workbook.save(file_path)
            except Exception as e:
                print(f"Erro ao salvar o arquivo: {e}")
            driver.quit()
            print("Final!")
            show_success_popup()
            break

        try:
            print("Salvando progresso...")
            workbook.save(file_path)
        except Exception as e:
            print(f"Erro ao salvar o arquivo: {e}")
        
        driver.get("http://psagapr01/saggestaoagu/pages/cadastro/profissional/consultar.xhtml")
        
        attempts = 0        
        while attempts < 5:  # Tentar até 5 vezes
            try:
                driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/div[1]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div/div/span/input").clear()
                driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/div[1]/div[2]/div[1]/div[1]/div[2]/div/div/div[1]/div/div/span/input").send_keys(siape)
                driver.execute_script("window.scrollTo(0, 800)")
                driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/div[1]/div[2]/div[2]/div/div[1]/button").click()
                wait = WebDriverWait(driver, 1000)
                time.sleep(3)
                # Usar JavaScript para verificar a presença do elemento
                element_present = driver.execute_script("""
                    return document.evaluate('/html/body/div[3]/div/div[2]', document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue !== null;
                """)
                if element_present:
                    alerta = driver.find_element(By.XPATH, "/html/body/div[3]/div/div[2]").text
                    if alerta == "Não foram encontrados registros para o conjunto de filtros informados. Revise os filtros e tente novamente.":
                        attempts += 1
                        if attempts >= 5:
                            linha += 1
                            break
                        continue             
                break
            
            except Exception as e:
                print(f"Erro ao tentar processar a matrícula: {e}")
                attempts += 1
                if attempts >= 5:
                    linha += 1  # Avança para a próxima linha após 5 tentativas
                    break  # Sai do loop interno e volta para o loop while principal
                continue  # Tenta novamente
        
        # Se o loop de tentativas foi concluído com sucesso (sem alertas), prossegue com o processamento normal
        if attempts < 5:
            try:
                wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div/form[1]/div[1]/div[2]/div[3]/div/div/div/div[1]/table/tbody/tr")))
                driver.execute_script("window.scrollTo(0, 800)")
            except:
                while True:
                    if driver.find_elements(By.XPATH, "/html/body/div[3]/div/form[1]/div[1]/div[2]/div[2]/div/div[1]/button"):
                        time.sleep(1)
                        try:
                            driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/div[1]/div[2]/div[2]/div/div[1]/button").click()
                            break
                        except:
                            time.sleep(1)
                            driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/div[1]/div[2]/div[2]/div/div[1]/button").click()
                            break
                    else:
                        time.sleep(0.1)
                        continue
            driver.execute_script("window.scrollTo(0, 800)")
                        
            time.sleep(2)

            
            # Verificar se o status é "Inativo"
            try:
                if driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/div[1]/div[2]/div[3]/div/div/div/div[1]/table/tbody/tr/td[6]").text == "Inativo":
                    time.sleep(1)
                    worksheet.cell(row=linha, column=10).value = "Inativo"
                    time.sleep(0.5)
                    linha += 1
                    continue
            except:
                print("Erro linha 94 - Stale Element - Repetindo processo")
                continue
            
            # Tentar clicar no link de alteração
            try:
                driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/div[1]/div[2]/div[3]/div/div/div/div[1]/table/tbody/tr/td[7]/span/span[2]/a[3]").click()
            except:
                try:
                    time.sleep(2)
                    driver.execute_script("document.getElementById('form:tabelaProfissionais:0:idAlterarCadastroProfissional').click();")
                except:
                    time.sleep(1)
                    worksheet.cell(row=linha, column=10).value = "Alteração do servidor encontra-se bloqueada para o seu perfil de acesso"
                    time.sleep(0.5)
                    print("Perfil bloqueado")
                    linha += 1
                    continue
            
            time.sleep(0.1)
            
            
            
            
            
            # Verificar mensagem de erro geral do sistema usando JavaScript
            error_present = driver.execute_script("""
                return document.evaluate(
                    '/html/body/div[3]/div/div[3]/div/ul/li/span',
                    document,
                    null,
                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                    null
                ).singleNodeValue !== null;
            """)

            if error_present:
                print("Erro geral do sistema. Contate o suporte técnico. Código do erro: [09421134511djEzMXAxNjUucHJldm5ldA&&]")
                time.sleep(0.5)
                linha += 1
                continue
            
            # Verificar se alguma área está selecionada
            area_elements = [
                driver.find_element(By.ID, "form:tipoAreaTrabalho:0").is_selected(),
                driver.find_element(By.ID, "form:tipoAreaTrabalho:1").is_selected(),
                driver.find_element(By.ID, "form:tipoAreaTrabalho:2").is_selected(),
                driver.find_element(By.ID, "form:tipoAreaTrabalho:3").is_selected()
            ]
            print(area_elements)
            
            if not any(area_elements):
                worksheet.cell(row=linha, column=10).value = "Cadastro do Servidor incompleto."
                print("Cadastro do Servidor incompleto.")
                linha += 1
                continue
            
            # Verificar presença do elemento usando JavaScript===================================
            element_present = driver.execute_script("""
                return document.evaluate(
                    '/html/body/div[3]/div/form[1]/span[4]/fieldset/div[2]/div[3]/div/div/div/div[2]/select',
                    document,
                    null,
                    XPathResult.FIRST_ORDERED_NODE_TYPE,
                    null
                ).singleNodeValue !== null;
            """)

            if element_present:
                try:
                    wait = WebDriverWait(driver, 10)
                    target_element = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div/form[1]/span[4]/fieldset/div[2]/div[3]/div/div/div/div[2]")))
                    select_element = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div/form[1]/span[4]/fieldset/div[2]/div[3]/div/div/div/div[2]/select")))
                    select = Select(select_element)
                    select.select_by_value("30")
                except TimeoutException:
                    print("Erro ao interagir com o elemento de paginação")
            else:
                print("Não possui paginação > avançando...")
            # Verificar presença do elemento usando JavaScript===================================

            
            # Verificar o número de linhas na tabela
            table_rows = driver.find_elements(By.XPATH, "/html/body/div[3]/div/form[1]/span[4]/fieldset/div[2]/div[3]/div/div/div/div/table/tbody/tr")
            
            # Se houver mais de 30 linhas, selecionar 30 linhas por página
            if len(table_rows) > 30:
                select.select_by_value("30")
            
            unidades = read_all_units(driver)
            print(unidades)
            
            # Verificar se a unidade está na lista
            if str(Unidade) in unidades:
                print(f"A unidade {Unidade} está na lista!")
            else:
                print(f"A unidade {Unidade} NÃO está na lista!")
                driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/span[4]/fieldset/div[2]/div[1]/div/div/div/input").send_keys(Unidade)
                driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/span[4]/fieldset/div[2]/div[1]/div/div/div/div[2]/button[1]").click()
            
            # Processar competências do profissional
            for cod in CodigoSv:
                cod = str(cod)
                #==nova solução para stale element===========================
                input_xpath = "html/body/div[3]/div/form[1]/span[5]/fieldset/div[2]/div/div[2]/span/div/div[1]/table/thead/tr/th[2]/div/input"
                for attempt in range(20):
                    try:
                        input_element = retry_find_element(driver, input_xpath)
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, input_xpath)))
                        time.sleep(1)
                        input_element.clear()
                        time.sleep(1)
                        input_element.send_keys(cod)
                        time.sleep(3)
                        break
                    except StaleElementReferenceException:
                        print(f"Tentativa {attempt + 1}: Elemento stale, tentando novamente...")
                        time.sleep(1)
                    except TimeoutException:
                        print(f"Tentativa {attempt + 1}: Timeout ao tentar encontrar o elemento, tentando novamente...")
                        time.sleep(1)
                    except Exception as e:
                        print(f"Tentativa {attempt + 1}: Ocorreu um erro: {e}")
                        time.sleep(1)
                    if attempt == 19:
                        time.sleep(1)
                        print("Não foi possível interagir com o elemento após várias tentativas.")
                        print("Reinicializando configuração do servidor...")
                        continue
                #==nova solução para stale element===========================
                
                
                
                # Esperar até que a tabela seja retornada====================
                if wait_for_table_update(driver, cod):
                    conteudo = driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/span[5]/fieldset/div[2]/div/div[2]/span/div/div[1]/table/tbody/tr/td[2]").text
                    wait = WebDriverWait(driver, 10)
                    codigo_procurado = str(cod)
                else:                    
                    try:
                        mensagem = driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/span[5]/fieldset/div[2]/div/div[2]/span/div/div[1]/table/tbody/tr/td").text
                        if mensagem == "Nenhum registro encontrado.":
                            worksheet.cell(row=linha, column=10).value = "Serviço não disponível para profissional diferente do tipo Administrativo."
                            time.sleep(0.5)
                            linha += 1
                            skip_to_next_iteration = True
                            break 
                    except:
                        print("A tabela não foi atualizada corretamente. Verifique o código ou aumente o tempo de espera.")
                # Esperar até que a tabela seja retornada====================
                
                
                
                
                max_retries = 3
                for attempt in range(max_retries):
                    try:
                        table = wait.until(EC.presence_of_element_located((By.ID, "form:tabelaServico_data")))
                        rows = table.find_elements(By.TAG_NAME, "tr")
                        
                        for i, row in enumerate(rows):
                            # Atualizar a referência dos elementos a cada tentativa
                            td_elements = row.find_elements(By.TAG_NAME, "td")
                            if len(td_elements) < 4:
                                continue
                            
                            codigo_celula = td_elements[1].text
                            if codigo_celula == codigo_procurado:
                                # Competência
                                checkbox_comp_id = f"form:tabelaServico:{i}:selecionarDeselecionarCompetencia"
                                if click_checkbox(driver, checkbox_comp_id, True):
                                    print(f"Checkbox de competência marcado para {cod}")
                                
                                # Atribuição (com tratamento específico)
                                checkbox_atrib_id = f"form:tabelaServico:{i}:selecionarDeselecionarAtribuicao"
                                for attr_attempt in range(3):
                                    try:
                                        if click_checkbox(driver, checkbox_atrib_id, AtribResp == "Sim"):
                                            status = "Marcado" if AtribResp == "Sim" else "Desmarcado"
                                            print(f"{status} atribuição na linha {i+1}")
                                        break
                                    except StaleElementReferenceException:
                                        print(f"Re-tentando atribuição ({attr_attempt+1}/3)")
                                        time.sleep(1)
                                
                                # Transferência (com reload de elementos)
                                checkbox_transf_id = f"form:tabelaServico:{i}:selecionarDeselecionarTransferencia"
                                if click_checkbox(driver, checkbox_transf_id, Trasf == "Sim"):
                                    status = "Marcado" if Trasf == "Sim" else "Desmarcado"
                                    print(f"{status} transferência na linha {i+1}")
                                
                                break
                        else:
                            print(f"Código {codigo_procurado} não encontrado.")
                        break  # Sai do loop de tentativas se bem sucedido
                        
                    except StaleElementReferenceException:
                        if attempt < max_retries - 1:
                            print(f"Tabela atualizada. Re-tentando ({attempt+1}/{max_retries})")
                            time.sleep(2)
                            # Atualizar referência da tabela
                            table = wait.until(EC.presence_of_element_located((By.ID, "form:tabelaServico_data")))
                            rows = table.find_elements(By.TAG_NAME, "tr")
                        else:
                            print(f"Falha após {max_retries} tentativas")
                            raise
                    except TimeoutException as e:
                        print(f"Timeout: {str(e)}")
                        if attempt < max_retries - 1:
                            print("Re-tentando após timeout...")
                            time.sleep(2)
                        else:
                            raise
            
            if skip_to_next_iteration:
                continue
            
            
            unidade = str(Unidade)
            if find_and_process_unit(driver, unidade):
                print("Abrindo Competencias do Profissional")
                
                def scroll_to_id(driver, element_id):
                    element = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, element_id))
                    )
                    driver.execute_script("arguments[0].scrollIntoView(true);", element)
                
                scroll_to_id(driver, "cmpModalCompetenciaServicoLocal:formPesquisaCompetencias:botaoFecharModalCompetenciaServicoLocal")
                time.sleep(1)
                verificar_e_bloquear_alteracoes(driver, BloquerAlteracoes)
                input_xpath = "/html/body/div[3]/div/form[5]/div/div/span/span/fieldset/div[2]/div/div/span/div/div[1]/table/thead/tr/th[2]/div/input"
                
                try:
                    for codigo in CodigoSv:
                        process_code(driver, codigo, AtivarMiExer, input_xpath)
                except ReiniciarProcessoException:
                    print("Reiniciando o processo principal...")
                    continue  # Reinicia o laço while principal
                
                # Confirmar do modal
                try:
                    wait = WebDriverWait(driver, 10)
                    botao_confirmar = wait.until(EC.element_to_be_clickable((By.ID, "cmpModalCompetenciaServicoLocal:formPesquisaCompetencias:botaoConfirmarModalCompetenciaServicoLocal")))
                    botao_confirmar.click()
                    print("Botão Confirmar clicado com sucesso.")
                    time.sleep(1)
                except Exception as e:
                    print(f"Erro ao clicar no botão Confirmar: {str(e)}")
            else:
                print("Falha no processamento da unidade")
            
            script = "window.scrollTo(0, document.body.scrollHeight);"
            driver.execute_script(script)
            
            # Confirmar e processar a alteração principal
            resultado = confirmar_e_processar(driver, worksheet, linha)
            if resultado:
                print("Servidor concluído com sucesso.")
            else:
                driver.get("http://psagapr01/saggestaoagu/pages/cadastro/profissional/consultar.xhtml")
                print("Houve um problema durante o processo.")
                update_status_func("Houve um problema durante o processo.")
                
            print("Servidor configurado com sucesso.")
            linha += 1

def run_automation_thread_saggestao(file_path, update_label_func=None, update_status_func=None):
    thread = threading.Thread(target=run_automation_saggestao, args=(file_path, update_label_func, update_status_func))
    thread.start()
    return thread
