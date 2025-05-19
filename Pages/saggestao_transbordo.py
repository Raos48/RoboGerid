from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, ElementNotInteractableException
from selenium.common.exceptions import JavascriptException, NoSuchElementException
from datetime import datetime
import openpyxl
import time
import tkinter as tk
from tkinter import messagebox
import subprocess
import threading
import queue
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
import time
import pyotp
from selenium.webdriver.support.ui import Select
import statistics
import traceback
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC



class ProcessoManager:
    def __init__(self):
        self.tempos_processamento = []
        self.total_registros = 0
        self.registros_processados = 0

    def inicializar_contagem(self, worksheet):
        self.total_registros = 0
        linha = 2
        while worksheet.cell(row=linha, column=1).value is not None:
            if worksheet.cell(row=linha, column=8).value is None:
                self.total_registros += 1
            linha += 1

    def registrar_tempo(self, tempo_inicio, tempo_fim):
        tempo_processamento = (tempo_fim - tempo_inicio).total_seconds()
        self.tempos_processamento.append(tempo_processamento)
        self.registros_processados += 1

    def calcular_tempo_estimado(self):
        if not self.tempos_processamento:
            return "Calculando..."
        
        tempo_medio = statistics.mean(self.tempos_processamento)
        registros_restantes = self.total_registros - self.registros_processados
        tempo_total_estimado = tempo_medio * registros_restantes
        
        horas = int(tempo_total_estimado // 3600)
        minutos = int((tempo_total_estimado % 3600) // 60)
        
        if horas > 0:
            return f"Tempo para conclusão: {horas}h {minutos}min"
        else:
            return f"Tempo para conclusão: {minutos}min"

    def get_status(self):
        tempo_estimado = self.calcular_tempo_estimado()
        registros_restantes = self.total_registros - self.registros_processados
        return f"Em execução... {tempo_estimado} | Registros restantes: {registros_restantes}"


class ReiniciarProcessoException(Exception):
    """Exceção personalizada para sinalizar a reinicialização do processo."""
    pass



def calcular_tempo_estimado(tempos_processamento, registros_restantes):
    if not tempos_processamento:
        return "Calculando..."
    
    tempo_medio = statistics.mean(tempos_processamento)
    tempo_total_estimado = tempo_medio * registros_restantes
    
    horas = int(tempo_total_estimado // 3600)
    minutos = int((tempo_total_estimado % 3600) // 60)
    
    if horas > 0:
        return f"Tempo estimado: {horas}h {minutos}min"
    else:
        return f"Tempo estimado: {minutos}min"


def show_success_popup():
    root = tk.Tk()
    root.withdraw() # Esconde a janela principal
    messagebox.showinfo("Sucesso", "A operação foi executada com sucesso!")
    root.destroy()

def close_workbook(workbook):
    try:
        workbook.close()
        print("Arquivo Excel fechado com sucesso.")
    except Exception as e:
        print(f"Erro ao fechar o arquivo Excel: {e}")



def get_codigo_sv(worksheet, linha):
    CodigoSv = []
    coluna = 11
    while True:
        valor = worksheet.cell(row=linha, column=coluna).value
        if valor is None:
            break
        CodigoSv.append(valor)
        coluna += 1
    return CodigoSv

def print_values(Unidade, CodServ, SiglaServ, UnidTrans, AreaMeio, GrupoMeio, Status):
    print("============================================================")
    print(f"Unidade:{Unidade}-Área Meio:{AreaMeio}-{GrupoMeio}")
    print(f"Serviço:", {CodServ}-{SiglaServ})
    print(f"Transferencia para:", {UnidTrans})    
    print(f"configurando...")

def click_checkbox(driver, checkbox_id, condition):
    max_attempts = 2
    for attempt in range(max_attempts):
        try:
            checkbox = driver.find_element(By.ID, checkbox_id)
            if checkbox.is_selected() != condition:
                checkbox.click()
                return True
            return False
        except StaleElementReferenceException:
            if attempt < max_attempts - 1:
                print(f"Stale element ao clicar no checkbox {checkbox_id}. Tentando novamente...")
                time.sleep(2)  # Espera antes de recarregar o elemento
                continue
            else:
                print(f"Falha após {max_attempts} tentativas no checkbox {checkbox_id}")
                return False
        except Exception as e:
            print(f"Erro inesperado ao clicar no checkbox {checkbox_id}: {str(e)}")
            return False
    return False


def run_automation_saggestao_transbordo(file_path, update_label_func=None, update_status_func=None, stop_event=None):
    
    processo_manager = ProcessoManager()
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    update_status_func("Iniciando execução.")


    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    processo_manager.inicializar_contagem(worksheet)


    driver.get("http://psagapr01/saggestaoagu/pages/cadastro/profissional/consultar.xhtml")
    driver.implicitly_wait(10)
    driver.maximize_window()
    
    #==== trecho exclusivo para RICARDO COMENTAR DEPOIS====================================
    # print(f"Aguardando Tela Código OTP")
    # WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.XPATH, "/html/body/div/div/main/div/div/div/div[2]/form/div/section[2]/input")))
    # print(f"Login realizado com sucesso.")
    # totp_key = "SCFW2ZRV7FZ4G3YO"
    # totp = pyotp.TOTP(totp_key)
    # codigo_totp = totp.now()
    # print(f"Código TOTP: {codigo_totp}")
    # wait = WebDriverWait(driver, 10)
    # driver.find_element(By.XPATH, "/html/body/div/div/main/div/div/div/div[2]/form/div/section[2]/input").click()
    # wait = WebDriverWait(driver, 10)
    # driver.find_element(By.XPATH, "/html/body/div/div/main/div/div/div/div[2]/form/div/section[2]/input").send_keys(codigo_totp)
    # wait = WebDriverWait(driver, 10)
    # driver.find_element(By.XPATH, "/html/body/div/div/main/div/div/div/div[2]/form/div/section[3]/input").click()
    # wait = WebDriverWait(driver, 10)
    # driver.find_element(By.XPATH, "/html/body/div/div[2]/button[3]").click()
    # wait = WebDriverWait(driver, 10)
    # driver.find_element(By.XPATH, "/html/body/div/div[3]/p[2]/a").click()
    # wait = WebDriverWait(driver, 10)
    # select_element = wait.until(EC.presence_of_element_located((By.ID, "domains")))
    # select = Select(select_element)
    # select.select_by_value("UO:01.001.PRES")
    # submit_button = driver.find_element(By.CSS_SELECTOR, "input[type='submit'][value='Enviar']")
    # submit_button.click()
    #===finalizar minha parte=======================
    
    print(f"Aguardando Login")
    WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div/form[1]/div[1]/div[2]/div[1]/div[1]/div[1]/div/span[2]")))
    print(f"Login realizado com sucesso.")
    linha = 2
    consecutive_errors = 0
    max_consecutive_errors = 3
    
    
    while True and not stop_event.is_set():
        
        if stop_event and stop_event.is_set():
            print("Interrupção solicitada, finalizando operação...")
            break
        
        try:            
            
            if stop_event and stop_event.is_set():
                break
            
            
            update_label_func(linha)
            update_status_func(processo_manager.get_status())
            
            Unidade = worksheet.cell(row=linha, column=1).value
            CodServ = worksheet.cell(row=linha, column=2).value
            SiglaServ = worksheet.cell(row=linha, column=3).value
            UnidTrans = worksheet.cell(row=linha, column=4).value
            Ativar = worksheet.cell(row=linha, column=5).value
            AreaMeio = worksheet.cell(row=linha, column=6).value
            GrupoMeio = worksheet.cell(row=linha, column=7).value
            Status = worksheet.cell(row=linha, column=8).value


            if Status is not None:
                linha += 1
                continue

            print_values(Unidade, CodServ, SiglaServ, UnidTrans, AreaMeio, GrupoMeio, Status)

            if not Unidade:
                try:
                    workbook.save(file_path)
                    print("Arquivo salvo com sucesso.")
                    close_workbook(workbook)
                except Exception as e:
                    print(f"Erro ao salvar ou fechar o arquivo: {e}")
                finally:
                    driver.quit()
                    print("Final!")
                    show_success_popup()
                    return

            try:
                print("Salvando progresso...")
                workbook.save(file_path)
            except Exception as e:
                print(f"Erro ao salvar o arquivo: {e}")  
                
                
            tempo_inicio = datetime.now()
            
            driver.get("https://psagapr01/saggestaoagu/pages/cadastro/unidade/consultar.xhtml")
            driver.execute_script("document.getElementById('formMenu:idGestaoUnidades').click();")        
            element = driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/fieldset/div[2]/span[1]/div[1]/div[1]/div/div/div/div/div[1]/input")
            if element.is_selected():
                pass
            else:
                driver.execute_script("document.getElementById('formConsultarUnidade:codigoTipoCodigo:0').click();") 


            input_xpath = "/html/body/div[3]/div/form[1]/fieldset/div[2]/span[1]/div[1]/div[1]/div/div/span/input"
            input_element = driver.find_element(By.XPATH, input_xpath)
            input_element.clear()
            time.sleep(1)
            input_element.send_keys(Unidade)
            time.sleep(1)
            driver.find_element(By.ID, "formConsultarUnidade:btnPesquisarUnidade").send_keys(Keys.ENTER)
            time.sleep(2)
            wait = WebDriverWait(driver, 1000)
            time.sleep(3)


            elements = driver.find_elements(By.XPATH, "/html/body/div[3]/div/div[2]/div/ul/li/span")            
            error_message = "Não foram encontrados registros para o conjunto de filtros informados. Revise os filtros e tente novamente."
            if elements and elements[0].text == error_message:
                consecutive_errors += 1
                print(f"Erro consecutivo {consecutive_errors}/{max_consecutive_errors}")
                
                if consecutive_errors >= max_consecutive_errors:
                    print(f"Erro persistente após {max_consecutive_errors} tentativas:")
                    print(elements[0].text)
                    worksheet.cell(row=linha, column=8).value = elements[0].text
                    linha += 1
                    consecutive_errors = 0  # Resetar o contador
                    continue
            else:
                consecutive_errors = 0  # Resetar o contador se não houver erro

            # Se chegou aqui, continua o processamento normal
            if elements and elements[0].text == error_message:
                # Tentativa única, não persistente
                print("Erro único detectado, tentando novamente...")
                continue


            try:
                elements = driver.find_elements(By.XPATH, "/html/body/div[3]/div/form[1]/fieldset/div[2]/div[3]/div/div/table/tbody/tr/td[5]")
                Situacao = elements[0].text
                if Situacao == "Inativa":
                    print(Situacao)
                    worksheet.cell(row=linha, column=8).value = Situacao                    
                    linha += 1                    
                    continue
            except:
                pass

            element_text = driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/fieldset/div[2]/div[3]/div/div/table/tbody/tr[1]/td[3]").text
            if element_text == "Unidade Externa":
                driver.execute_script("document.getElementById('formConsultarUnidade:tableUnidades:1:btnConfigurarDetalhar').click();")
            else:
                driver.execute_script("document.getElementById('formConsultarUnidade:tableUnidades:0:btnConfigurarDetalhar').click();")
            time.sleep(2)
            while True:
                time.sleep(0.2)  # Espera de 200 milissegundos
                if driver.find_elements(By.ID, "formManterUO"):
                    break        
            driver.execute_script("document.getElementById('formManterUO:panelPaisesAcordoInternacional').scrollIntoView();")
            #time.sleep(1.5)
            #WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, "panelConfirmarSelecaoGrupoAreaMeio:frmConfirmarSelecaoGrupoAreaMeio:dlgConfirmarSelecaoGrupoAreaMeio_modal")))   
            #driver.execute_script("document.getElementById('panelConfirmarSelecaoGrupoAreaMeio:frmConfirmarSelecaoGrupoAreaMeio:btnAjaxConfirmarConfirmarSelecaoGrupoAreaMeio').click();")
            #time.sleep(1.5)
            if AreaMeio == "Sim":
                #time.sleep(1.5)
                driver.execute_script("document.getElementById('formManterUO:grupoServicoareaFim:0').click();")
                time.sleep(1.5)
                driver.execute_script("document.getElementById('panelConfirmarSelecaoUnidadeAreaMeioFim:frmConfirmarSelecaoUnidadeAreaMeioFim:btnAjaxConfirmarConfirmarSelecaoUnidadeAreaMeioFim').click();");
                time.sleep(1.5)
                driver.execute_script("document.getElementById('formManterUO:panelExecucoes').scrollIntoView();")
                time.sleep(1)
                driver.find_element(By.XPATH, "/html/body/div[3]/div/form[1]/span[1]/fieldset/div[2]/div/div[2]/span/div/div/div/label").click()
                time.sleep(1)
                try:        
                    driver.find_element(By.XPATH, f"//ul[@id='formManterUO:selectGrupoServico_items']//li[@data-label='{GrupoMeio}']").click()
                except Exception as e:
                    print(f"Erro ao selecionar opção: {str(e)}")
                time.sleep(1.5)
                driver.execute_script("document.getElementById('panelConfirmarSelecaoGrupoAreaMeio:frmConfirmarSelecaoGrupoAreaMeio:btnAjaxConfirmarConfirmarSelecaoGrupoAreaMeio').click();")

            time.sleep(2)
            driver.execute_script("document.getElementById('formManterUO:tabelaPaisesAcordo:nomePaisAcordo').scrollIntoView();")
            time.sleep(0.5)
            
            input_xpath_1 = "/html/body/div[3]/div/form[1]/span[1]/fieldset/div[2]/div/div[3]/div[1]/table/thead/tr/th[2]/div/input"
            driver.find_element(By.XPATH, input_xpath_1).clear()
            time.sleep(0.5)
            driver.find_element(By.XPATH, input_xpath_1).send_keys(CodServ)
            time.sleep(2)
            
            xpath = "/html/body/div[3]/div/form[1]/span[1]/fieldset/div[2]/div/div[3]/div[1]/table/tbody/tr[1]/td[9]/input"
            wait = WebDriverWait(driver, 10)
            element = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
            
            input_xpath_2 = "/html/body/div[3]/div/form[1]/span[1]/fieldset/div[2]/div/div[3]/div[1]/table/thead/tr/th[3]/div/input"
            driver.find_element(By.XPATH, input_xpath_2).clear()
            time.sleep(0.5)
            driver.find_element(By.XPATH, input_xpath_2).send_keys(SiglaServ)
            time.sleep(2)
            
            
            
            xpath = "/html/body/div[3]/div/form[1]/span[1]/fieldset/div[2]/div/div[3]/div[1]/table/tbody/tr[1]/td[9]/input"
            max_attempts = 5
            attempts = 0
            while attempts < max_attempts:
                try:
                    wait = WebDriverWait(driver, 10)
                    element = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
                    
                    if Ativar == "Sim":
                        if not element.is_selected():
                            element.click()
                    elif Ativar == "Não":
                        if element.is_selected():
                            element.click()
                    
                    # Se chegou aqui, conseguiu completar a operação
                    break
                    
                except StaleElementReferenceException:
                    attempts += 1
                    print(f"Elemento obsoleto. Tentativa {attempts} de {max_attempts}")
                    time.sleep(1)  # Pequena pausa antes de tentar novamente
                    
                    if attempts == max_attempts:
                        print("Não foi possível manipular o checkbox após várias tentativas")
                        raise  # Relanço a exceção se esgotar todas as tentativas
            
            
            time.sleep(1)


            if Ativar == "Sim" and UnidTrans:
                driver.execute_script("document.getElementById('formManterUO:tabelaServicosExecutados:0:btnConfigurarServico').click();")
                time.sleep(1.5)
                element = driver.find_element(By.XPATH, "/html/body/div[3]/div/form[6]/div/div[2]/span/fieldset/div[2]/div[3]/div[1]/div/div/div[3]/input")
                if element.is_selected():
                    time.sleep(1)
                else:    
                    if UnidTrans is None:
                        driver.execute_script("document.getElementById('panelModalConfigurarServicoPrestado:formModalConfigurarServicoPrestado:fluxoInicialTarefa:1').click();")
                    else:
                        driver.execute_script("document.getElementById('panelModalConfigurarServicoPrestado:formModalConfigurarServicoPrestado:fluxoInicialTarefa:2').click();")
                        time.sleep(1)        
                        driver.execute_script("document.getElementById('panelModalConfigurarServicoPrestado:formModalConfigurarServicoPrestado:btnPesquisarUnidadeFluxoTarefa').click();")
                        time.sleep(1)
                        driver.find_element(By.XPATH, "/html/body/div[3]/div/form[9]/div/div/span/fieldset/div[2]/div[1]/div[1]/div/div/span/input").send_keys(UnidTrans)
                        time.sleep(1)
                        driver.execute_script("document.getElementById('panelModalConfigurarFluxoInicialTarefa:panelModalConsultarUnidadeFluxoInicialTarefa:formModalConsultarUnidade:btnPesquisarUnidadeModal').click();")
                        time.sleep(3)
                        driver.execute_script("document.getElementById('panelModalConfigurarFluxoInicialTarefa:panelModalConsultarUnidadeFluxoInicialTarefa:formModalConsultarUnidade:tableUnidadesModal:0:selecionarUnidade').click();")
                        time.sleep(1)
                        driver.execute_script("document.getElementById('panelModalConfigurarServicoPrestado:formModalConfigurarServicoPrestado:btnConfirmarConfigurarServicoPrestado').click();")
                        time.sleep(3)
                        elements = driver.find_elements(By.XPATH, "/html/body/div[3]/div/form[4]/div/div[2]/div[1]/div/ul/li/span")
                        if len(elements) == 1:
                            element_text = elements[0].text    
                            if element_text == "Os tempos do serviço só podem ser alterados para valores menores ou iguais aos valores de referência.":
                                time.sleep(0.5)  # Pausa de 500ms
                                input_element = driver.find_element(By.XPATH, "/html/body/div[3]/div/form[4]/div/div[2]/span/fieldset/div[2]/div[2]/div[4]/div/input")
                                input_element.clear()
                                time.sleep(1)
                                reference_text = driver.find_element(By.XPATH, "/html/body/div[3]/div/form[4]/div/div[2]/span/fieldset/div[2]/div[2]/div[3]/div/div/span").text
                                input_element.send_keys(reference_text)
                                time.sleep(1)
                                driver.execute_script("document.getElementById('panelModalConfigurarServicoPrestado:formModalConfigurarServicoPrestado:btnConfirmarConfigurarServicoPrestado').click();")
                                time.sleep(2)
                        driver.execute_script("document.getElementById('formManterUO:botaoConfirmarAlterarUO').click();")
                        time.sleep(2)
                        print(driver.find_element(By.XPATH, "/html/body/div[3]/div/div[2]/div").text)

            # elements = driver.find_elements(By.XPATH, "/html/body/div[3]/div/form[4]/div/div[2]/div[2]/div/ul/li/span/label")
            # if len(elements) > 0:
            #     element_text = elements[0].text
            #     if element_text == "A tarefa não pode ser configurada porque o perfil do usuário é menor do que o da alteração anterior.":
            #         planilha.cell(row=linha, column=8).value = element_text
            #         driver.execute_script("document.getElementById('panelModalConfigurarServicoPrestado:formModalConfigurarServicoPrestado:btnCancelarConfigurarServicoPrestado').click();")
            #         linha += 1
            #         #continue 

            driver.execute_script("document.getElementById('formManterUO:botaoConfirmarAlterarUO').scrollIntoView();")
            time.sleep(1)
            driver.execute_script("document.getElementById('formManterUO:botaoConfirmarAlterarUO').click();")
            time.sleep(1)
            print(driver.find_element(By.XPATH, "/html/body/div[3]/div/div[2]/div").text)    
            worksheet.cell(row=linha, column=8).value = driver.find_element(By.XPATH, "/html/body/div[3]/div/div[2]/div").text
            
            #cronometro fim
            tempo_fim = datetime.now()
            processo_manager.registrar_tempo(tempo_inicio, tempo_fim)            
            update_status_func(processo_manager.get_status())            
            linha += 1        
            continue
        except Exception as e:
            
            if stop_event and stop_event.is_set():
                print("Operação interrompida durante tratamento de erro")
                break            
            tb_str = traceback.format_exc()           
            print(f"Erro inesperado: {str(e)}. Detalhes:\n{tb_str}. Reiniciando o loop.")
            continue
        # Adicione o bloco finally para limpeza
        finally:
            if stop_event and stop_event.is_set():
                print("Executando limpeza pós-interrupção...")
                try:
                    driver.quit()
                    if update_status_func:
                        update_status_func("Operação interrompida pelo usuário")
                    workbook.save(file_path)
                    close_workbook(workbook)
                except Exception as e:
                    print(f"Erro durante a limpeza: {e}")

def run_automation_thread_saggestao_transbordo(file_path, update_label_func=None, update_status_func=None, stop_event=None):
    thread = threading.Thread(target=run_automation_saggestao_transbordo, args=(file_path, update_label_func, update_status_func,stop_event))
    thread.start()
    return thread


